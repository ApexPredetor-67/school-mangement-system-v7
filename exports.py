from io import BytesIO
from pathlib import Path
import json, base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import mm


def build_xlsx(rows, title='School Report'):
    wb = Workbook(); ws = wb.active; ws.title = 'Report'
    ws.append([title]); ws.append([])
    if rows:
        headers = list(rows[0].keys()); ws.append(headers)
        for r in rows: ws.append([r.get(h) for h in headers])
        for c in ws[3]: c.font = Font(bold=True); c.fill = PatternFill('solid', fgColor='EAF0FF')
        for col in ws.columns:
            width = max(len(str(x.value or '')) for x in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(width, 35)
    out = BytesIO(); wb.save(out); out.seek(0); return out


def build_pdf(rows, title='School Report', subtitle=''):
    out = BytesIO(); doc = SimpleDocTemplate(out, pagesize=(A4[1], A4[0]), rightMargin=18,leftMargin=18,topMargin=18,bottomMargin=18)
    styles = getSampleStyleSheet(); story=[Paragraph(title, styles['Title'])]
    if subtitle: story.append(Paragraph(subtitle, styles['Normal']))
    story.append(Spacer(1,8))
    if rows:
        headers=list(rows[0].keys()); data=[headers]+[[str(r.get(h,'')) for h in headers] for r in rows]
        t=Table(data, repeatRows=1)
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#EAF0FF')),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#D7DCE6')),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('FONTSIZE',(0,0),(-1,-1),8)]))
        story.append(t)
    doc.build(story); out.seek(0); return out


def _grade(pct):
    if pct >= 91: return 'A1'
    if pct >= 81: return 'A2'
    if pct >= 71: return 'B1'
    if pct >= 61: return 'B2'
    if pct >= 51: return 'C1'
    if pct >= 41: return 'C2'
    if pct >= 33: return 'D'
    return 'E'


def _decode_signature(data, width=34*mm, height=14*mm):
    if not data or not str(data).startswith('data:image'):
        return None
    try:
        raw = base64.b64decode(str(data).split(',',1)[1])
        return Image(BytesIO(raw), width=width, height=height, preserveAspectRatio=True, mask='auto')
    except Exception:
        return None


def _header_story(story, student, session_name, logo_path=None):
    styles = getSampleStyleSheet()
    title = ParagraphStyle('cardtitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=19, leading=21, alignment=TA_CENTER, textColor=colors.HexColor('#173A63'), spaceAfter=2)
    sub = ParagraphStyle('cardsmall', parent=styles['Normal'], fontName='Helvetica', fontSize=8.5, leading=10, alignment=TA_CENTER, textColor=colors.HexColor('#58667A'))
    brand = []
    if logo_path:
        try: brand.append(Image(str(logo_path), width=22*mm, height=22*mm, preserveAspectRatio=True, mask='auto'))
        except Exception: pass
    brand.extend([Paragraph('D.A.V. PUBLIC SCHOOL', title), Paragraph('ACADEMIC PROGRESS REPORT', sub), Paragraph(f'Academic Session: <b>{session_name}</b>', sub)])
    if len(brand) == 4:
        t = Table([[brand[0], brand[1], brand[2], brand[3]]], colWidths=[26*mm, 55*mm, 58*mm, 38*mm])
    else:
        t = Table([[Paragraph('D.A.V. PUBLIC SCHOOL', title), Paragraph('ACADEMIC PROGRESS REPORT', sub), Paragraph(f'Academic Session: <b>{session_name}</b>', sub)]], colWidths=[65*mm, 65*mm, 55*mm])
    t.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(-1,-1),'CENTER'),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
    story.append(t); story.append(Spacer(1,4))


def _boxed_table(data, widths, header_rows=1, fontsize=7.5, header_bg='#EAF0FF'):
    t=Table(data,colWidths=widths,repeatRows=header_rows)
    st=[('GRID',(0,0),(-1,-1),0.35,colors.HexColor('#C9D2DE')),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('FONTSIZE',(0,0),(-1,-1),fontsize),('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4)]
    if header_rows:
        st += [('BACKGROUND',(0,0),(-1,header_rows-1),colors.HexColor(header_bg)),('FONTNAME',(0,0),(-1,header_rows-1),'Helvetica-Bold')]
    t.setStyle(TableStyle(st)); return t


def build_report_card(student, session_name, attendance, term1, term2, subjects, exams, marks_by_subject, teacher_signature=None, config=None, logo_path=None, assessment_by_subject=None):
    """Generate an original, configurable school report card. It is not a copy of the photographed paper."""
    config = config or {}
    co = config.get('co_scholastic', {})
    dis = config.get('discipline', {})
    health = config.get('health', {})
    layout = config.get('layout', {})
    report_title=layout.get('report_title','STUDENT PROGRESS REPORT')
    attendance_title=layout.get('attendance_title','ATTENDANCE SUMMARY')
    scholastic_title=layout.get('scholastic_title','SCHOLASTIC PERFORMANCE')
    development_title=layout.get('development_title','CO-SCHOLASTIC & PERSONAL DEVELOPMENT')
    guide_title=layout.get('guide_title','PROGRESS REPORT GUIDE')
    class_details_title=layout.get('class_details_title','CLASS DETAILS')
    pass_rule=layout.get('pass_rule','33% in every subject.')
    teacher_signature_label=layout.get('teacher_signature_label','Class Teacher Signature')
    principal_signature_label=layout.get('principal_signature_label','Principal Signature')
    parent_signature_label=layout.get('parent_signature_label','Parent Signature')
    assessment_by_subject = assessment_by_subject or {}
    out=BytesIO(); doc=SimpleDocTemplate(out,pagesize=A4,rightMargin=11*mm,leftMargin=11*mm,topMargin=10*mm,bottomMargin=10*mm)
    styles=getSampleStyleSheet(); body=ParagraphStyle('body2',parent=styles['BodyText'],fontSize=8.2,leading=10.2,textColor=colors.HexColor('#26344A')); small=ParagraphStyle('sm2',parent=body,fontSize=7.3,leading=9); h=ParagraphStyle('h2x',parent=styles['Heading2'],fontSize=11.5,leading=13,textColor=colors.HexColor('#173A63'),spaceBefore=3,spaceAfter=6); center=ParagraphStyle('center',parent=body,alignment=TA_CENTER)
    story=[]

    # PAGE 1 — identity + attendance + school summary
    _header_story(story,student,session_name,logo_path)
    story.append(Paragraph(report_title,h))
    photo = None
    if getattr(student,'account',None) and getattr(student.account,'profile_picture_path',None):
        p=student.account.profile_picture_path
        if p:
            try:
                pth=Path(str(p))
                if not pth.is_absolute(): pth=Path(__file__).resolve().parent/'static'/str(p).replace('static/','',1)
                photo=Image(str(pth),width=26*mm,height=33*mm,preserveAspectRatio=True,mask='auto')
            except Exception: pass
    house=config.get('house','')
    info=[
      ['Student Name', student.name or '', 'Class', student.class_name or '', 'Section', student.section or ''],
      ['Admission No.', student.admission_number or '', 'Roll No.', student.roll_number or '', 'House', house],
      ['Date of Birth', student.date_of_birth.strftime('%d.%m.%Y') if student.date_of_birth else '', "Mother's Name", student.mother_name or '', "Father's Name", student.father_name or ''],
      ['Contact No.', student.phone or '', 'Previous School', student.previous_school or '', 'Address', student.address or ''],
    ]
    t=_boxed_table(info,[28*mm,38*mm,20*mm,38*mm,23*mm,37*mm],0,7.3)
    if photo:
        wrap=Table([[t,photo]],colWidths=[154*mm,31*mm]); wrap.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),('BOX',(1,0),(1,0),0.4,colors.HexColor('#C9D2DE')),('ALIGN',(1,0),(1,0),'CENTER')]))
        story.append(wrap)
    else: story.append(t)
    story.append(Spacer(1,6)); story.append(Paragraph(attendance_title,h))
    att_rows=[['Period','Working Days','Present','Absent','Attendance %'],['Term I',term1.get('working_days',0),term1.get('present',0),term1.get('absent',0),f"{term1.get('percentage',0):.2f}%"],['Term II',term2.get('working_days',0),term2.get('present',0),term2.get('absent',0),f"{term2.get('percentage',0):.2f}%"],['Whole Year / Current',attendance.get('working_days',0),attendance.get('present',0),attendance.get('absent',0),f"{attendance.get('percentage',0):.2f}%"]]
    story.append(_boxed_table(att_rows,[45*mm,31*mm,28*mm,28*mm,37*mm],1,7.6)); story.append(Spacer(1,6))
    story.append(Paragraph(class_details_title,h));
    class_rows=[['Class Teacher',config.get('class_teacher_name','')],['Date & Result',config.get('date_result','')],['Remarks',config.get('remarks','') or '']]
    story.append(_boxed_table(class_rows,[40*mm,129*mm],0,7.5)); story.append(PageBreak())

    # PAGE 2 — scholastic
    _header_story(story,student,session_name,logo_path); story.append(Paragraph(scholastic_title,h)); story.append(Paragraph('Periodic tests and final examination are recorded independently. The best-two PT contribution is normalized to 5 marks and can be combined with assessment components to form the internal assessment /20.',small)); story.append(Spacer(1,5))
    pt_exams=[e for e in exams if str(e.name).upper() in {'PT-1','PT-2','PT-3'}]; final=[e for e in exams if e.is_final or str(e.name).lower().startswith('final')]
    headers=['Subject']+[e.name+' /'+str(e.max_marks) for e in pt_exams]+['Best 2 PTs /5','Final /80','Internal /20','Total /100','Grade']
    schol=[headers]
    total_marks=0; total_max=0
    for subj in subjects:
        row=[subj.name]; pts=[]
        for e in pt_exams:
            m=marks_by_subject.get((subj.code,e.id)); row.append(f"{m.marks:g}" if m and m.marks is not None else '—');
            if m and m.marks is not None: pts.append((m.marks/m.max_marks)*40)
        pts_sorted=sorted(pts,reverse=True)[:2]; best5=(sum(pts_sorted)/len(pts_sorted))/8 if pts_sorted else 0
        comp=assessment_by_subject.get(subj.code,{})
        ma=float(comp.get('multiple_assessment',0) or 0); se=float(comp.get('subject_enrichment',0) or 0); po=float(comp.get('portfolio',0) or 0)
        internal=min(20,best5+ma+se+po)
        final_mark='—'; final_val=0
        for e in final:
            m=marks_by_subject.get((subj.code,e.id));
            if m and m.marks is not None: final_mark=f'{m.marks:g}'; final_val=m.marks; break
        total=internal+final_val; pct=total if (final and final[0].max_marks==80) else total
        if final_mark!='—': total_marks += total; total_max += 100
        row += [f'{best5:.1f}',final_mark,f'{internal:.1f}',f'{total:.1f}' if final_mark!='—' else '—',_grade(total) if final_mark!='—' else '—']; schol.append(row)
    widths=[40*mm]+[17*mm]*len(pt_exams)+[22*mm,20*mm,22*mm,22*mm,16*mm]
    story.append(_boxed_table(schol,widths,1,6.25));
    overall_pct=(total_marks/total_max*100) if total_max else 0
    story.append(Spacer(1,6)); story.append(Paragraph(f'<b>Total Marks Obtained:</b> {total_marks:.1f} &nbsp;&nbsp; <b>Total Max:</b> {total_max} &nbsp;&nbsp; <b>Percentage:</b> {overall_pct:.2f}% &nbsp;&nbsp; <b>Overall Grade:</b> {_grade(overall_pct)}',body))
    story.append(PageBreak())

    # PAGE 3 — co-scholastic, discipline, health, signatures
    _header_story(story,student,session_name,logo_path); story.append(Paragraph(development_title,h))
    co_rows=[['Activity','Grade'],['Work Education',co.get('work_education','')],['Art Education (Visual & Performing Arts)',co.get('art_education','')],['Health & Physical Education',co.get('health_physical','')]]
    story.append(_boxed_table(co_rows,[115*mm,25*mm],1,7.6)); story.append(Spacer(1,6))
    dis_rows=[['Element','Grade'],['Discipline',dis.get('discipline','')],['Regularity',dis.get('regularity','')],['Punctuality',dis.get('punctuality','')]]
    story.append(_boxed_table(dis_rows,[115*mm,25*mm],1,7.6)); story.append(Spacer(1,6))
    h_rows=[['Health Status','Height','Weight'],['Term I',health.get('term1_height',''),health.get('term1_weight','')],['Term II',health.get('term2_height',''),health.get('term2_weight','')]]
    story.append(_boxed_table(h_rows,[75*mm,35*mm,30*mm],1,7.6)); story.append(Spacer(1,8))
    story.append(Paragraph('REMARKS',h)); story.append(Paragraph((config.get('remarks','') or '—').replace('\n','<br/>'),body)); story.append(Spacer(1,8)); story.append(Paragraph('PRINCIPAL\'S REMARKS',h)); story.append(Paragraph((config.get('principal_remarks','') or '—').replace('\n','<br/>'),body)); story.append(Spacer(1,8))
    next_session=config.get('next_academic_session',''); start=config.get('session_begins',''); summer=config.get('summer_break_from',''); reopen=config.get('school_reopens','')
    date_rows=[['Next Academic Session',next_session],['Session Begins',start],['Summer Break From',summer],['School Re-opens',reopen]]
    story.append(_boxed_table(date_rows,[55*mm,105*mm],0,7.4)); story.append(Spacer(1,10))
    sig_img=_decode_signature(teacher_signature)
    sig_table=[[sig_img or '', '', ''],[teacher_signature_label,principal_signature_label,parent_signature_label]]
    st=Table(sig_table,colWidths=[55*mm,55*mm,55*mm],rowHeights=[18*mm,6*mm]); st.setStyle(TableStyle([('LINEBELOW',(0,0),(-1,0),0.5,colors.HexColor('#AAB4C1')),('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('FONTSIZE',(0,1),(-1,1),7.5),('TEXTCOLOR',(0,1),(-1,1),colors.HexColor('#58667A'))])); story.append(st); story.append(PageBreak())

    # PAGE 4 — grading/rules
    _header_story(story,student,session_name,logo_path); story.append(Paragraph(guide_title,h)); rules=[['Rule','School Policy'],['Pass percentage',pass_rule],['Result integrity','Declared results are locked and are not altered through normal user workflows.'],['Assessment basis','Performance, regularity, participation, class work and co-curricular development may be considered for teacher remarks.'],['Re-test','No re-test is assumed unless the school config explicitly provides one.'],['Principal authority','Final administrative decision rests with the Principal.']]
    story.append(_boxed_table(rules,[48*mm,112*mm],1,7.5)); story.append(Spacer(1,8)); story.append(Paragraph('8-POINT SCHOLASTIC GRADING SCALE',h)); g=[['Marks Range','Grade'],['91–100','A1'],['81–90','A2'],['71–80','B1'],['61–70','B2'],['51–60','C1'],['41–50','C2'],['33–40','D'],['Below 33%','E']]; story.append(_boxed_table(g,[65*mm,35*mm],1,8)); story.append(Spacer(1,8)); story.append(Paragraph('5-POINT CO-SCHOLASTIC SCALE',h)); c=[['Meaning','Grade'],['Highly Competent','A'],['Quite Capable','B'],['Performs Satisfactorily','C'],['Trying Well','D'],['Can Do Better','E']]; story.append(_boxed_table(c,[65*mm,35*mm],1,8)); story.append(Spacer(1,14)); story.append(Paragraph('This report card was generated from the school management system using the configured academic record and teacher-entered report-card details.',small))
    doc.build(story); out.seek(0); return out

