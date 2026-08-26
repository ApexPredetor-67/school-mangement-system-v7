from datetime import date, datetime, timedelta, time as dt_time
from functools import wraps
from io import BytesIO
import hashlib, ipaddress, json, os, secrets, io, re, hmac, base64
from pathlib import Path
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
from flask import Flask, abort, flash, jsonify, redirect, render_template, request, send_file, session, url_for, g
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash
from sqlalchemy import func, or_, and_, text

try:
    import numpy as np
    import cv2
except Exception:
    np = None
    cv2 = None

from models import db, Account, Student, Parent, ParentStudent, Teacher, TeacherAssignment, TeacherSubjectAssignment, Subject, Exam, Mark, AssessmentComponent, Attendance, SchoolCalendar, Announcement, ResultPublication, AuditEvent, SchoolSetting, SchoolClock, PublishedReport, FeeInvoice, FeePayment, FeeStructure, FeeStructureDocument, FeePaymentWindow, ReportCardConfig
from exports import build_xlsx, build_pdf, build_report_card
from face_utils import available as face_available, encode_frame, best_match, match_distance

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

try:
    import pytesseract
except Exception:
    pytesseract = None

load_dotenv()
BASE_DIR = Path(__file__).resolve().parent
app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.getenv('SECRET_KEY', secrets.token_hex(32)),
    SQLALCHEMY_DATABASE_URI=os.getenv('DATABASE_URL','sqlite:///'+str(BASE_DIR/'school.db')).replace('postgres://','postgresql://'),
    SQLALCHEMY_TRACK_MODIFICATIONS=False,
    SQLALCHEMY_ENGINE_OPTIONS={
        'pool_pre_ping': True,
        'pool_recycle': 280,
        'pool_size': int(os.getenv('DB_POOL_SIZE','2')),
        'max_overflow': int(os.getenv('DB_MAX_OVERFLOW','2')),
        'pool_timeout': int(os.getenv('DB_POOL_TIMEOUT','10')),
    },
    JWT_SECRET_KEY=os.getenv('JWT_SECRET_KEY') or os.getenv('SECRET_KEY') or secrets.token_hex(32),
    JWT_ACCESS_TTL_MINUTES=int(os.getenv('JWT_ACCESS_TTL_MINUTES','60')),
    APP_TIMEZONE=os.getenv('APP_TIMEZONE','Asia/Kolkata'),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=os.getenv('SESSION_COOKIE_SECURE','false').lower()=='true',
)
if os.getenv('TRUST_PROXY_HEADERS','false').lower()=='true':
    # Render terminates TLS at its load balancer and forwards the original client IP.
    # Only enable this when the app is actually behind a trusted single proxy hop.
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
db.init_app(app)


def csv_values(name):
    return [x.strip() for x in os.getenv(name,'').split(',') if x.strip()]

def client_ip():
    return request.remote_addr or 'unknown'

def ip_allowed(name):
    allowed = csv_values(name)
    if not allowed: return True
    ip = ipaddress.ip_address(client_ip())
    for item in allowed:
        try:
            if ip == ipaddress.ip_address(item): return True
        except ValueError:
            try:
                if ip in ipaddress.ip_network(item, strict=False): return True
            except ValueError: pass
    return False


def _jwt_payload_from_request():
    auth=request.headers.get('Authorization','')
    if not auth.lower().startswith('bearer '):
        return None
    token=auth.split(None,1)[1].strip()
    if not token:
        return None
    try:
        import jwt
        return jwt.decode(token, app.config['JWT_SECRET_KEY'], algorithms=['HS256'])
    except Exception:
        return None


def jwt_account_from_request():
    payload=_jwt_payload_from_request()
    if not payload:
        return None
    try:
        aid=int(payload.get('sub'))
    except (TypeError, ValueError):
        return None
    acct=db.session.get(Account, aid)
    if not acct or not acct.active:
        return None
    g.jwt_authenticated=True
    return acct


def _issue_jwt(acct):
    import jwt
    now=datetime.utcnow()
    exp=now+timedelta(minutes=app.config['JWT_ACCESS_TTL_MINUTES'])
    return jwt.encode({'sub':str(acct.id),'role':acct.role,'iat':int(now.timestamp()),'exp':int(exp.timestamp())}, app.config['JWT_SECRET_KEY'], algorithm='HS256')


@app.before_request
def _csrf_guard():
    # Anonymous pages still receive a token so public POST forms are protected.
    if not session.get('csrf'):
        session['csrf']=secrets.token_urlsafe(32)
    if request.method not in {'POST','PUT','PATCH','DELETE'}:
        return None
    if request.path in {'/login','/api/auth/login','/healthz'}:
        return None
    # Bearer-token API calls are authenticated independently of browser CSRF.
    if request.headers.get('Authorization','').lower().startswith('bearer '):
        return None
    supplied=request.headers.get('X-CSRF-Token') or request.form.get('_csrf')
    expected=session.get('csrf','')
    if not supplied or not expected or not hmac.compare_digest(str(supplied),str(expected)):
        if request.path.startswith('/api/'):
            return jsonify({'error':'Invalid or missing CSRF token.'}),400
        return ('Invalid or missing CSRF token.',400)
    return None


def staff_network_required():
    def deco(fn):
        @wraps(fn)
        def wrapper(*args,**kwargs):
            if not ip_allowed('STAFF_ALLOWED_IPS'):
                return ('<h1>403</h1><p>Staff services are available only from an authorized school network.</p>',403)
            return fn(*args,**kwargs)
        return wrapper
    return deco

def audit_network_required():
    def deco(fn):
        @wraps(fn)
        def wrapper(*args,**kwargs):
            if not ip_allowed('AUDIT_LOG_ALLOWED_IPS'):
                return ('<h1>403</h1><p>Restricted audit service.</p>',403)
            return fn(*args,**kwargs)
        return wrapper
    return deco

def login_required(role=None):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args,**kwargs):
            acct=None
            if request.headers.get('Authorization','').lower().startswith('bearer '):
                acct=jwt_account_from_request()
            if not acct:
                aid=session.get('account_id')
                if aid:
                    acct=db.session.get(Account, aid)
            if not acct or not acct.active:
                if request.path.startswith('/api/'):
                    return jsonify({'error':'Authentication required.'}),401
                return redirect(url_for('login', next=request.path))
            g.current_account=acct
            if role and acct.role != role: abort(403)
            if acct.must_change_password and request.endpoint not in {'change_credentials','logout'} and not g.get('jwt_authenticated'):
                return redirect(url_for('change_credentials'))
            return fn(*args,**kwargs)
        return wrapper
    return deco

def staff_required(fn):
    @wraps(fn)
    @staff_network_required()
    @login_required()
    def wrapper(*args,**kwargs):
        acct=current_account();
        if not acct or acct.role not in {'admin','teacher'}: abort(403)
        return fn(*args,**kwargs)
    return wrapper

def admin_required(fn):
    @wraps(fn)
    @staff_network_required()
    @login_required('admin')
    def wrapper(*args,**kwargs): return fn(*args,**kwargs)
    return wrapper

def current_account():
    acct=getattr(g,'current_account',None)
    if acct: return acct
    aid=session.get('account_id')
    return db.session.get(Account, aid) if aid else None


def log_audit(action, target_type='', target_id='', extra=None):
    acct=current_account(); actor=acct.username if acct else session.get('audit_username','system'); role=acct.role if acct else 'audit'
    prev=db.session.query(AuditEvent).order_by(AuditEvent.id.desc()).first()
    prev_hash=prev.event_hash if prev else ''
    payload={'action':action,'target_type':target_type,'target_id':str(target_id or ''),'ip':client_ip(),'at':datetime.utcnow().isoformat(),'extra':extra or {}}
    raw=prev_hash+json.dumps(payload,sort_keys=True,separators=(',',':'))
    h=hashlib.sha256(raw.encode()).hexdigest()
    db.session.add(AuditEvent(actor_username=actor,actor_role=role,action=action,target_type=target_type,target_id=str(target_id or ''),ip_address=client_ip(),user_agent=request.headers.get('User-Agent',''),metadata_json=json.dumps(extra or {}),previous_hash=prev_hash,event_hash=h))


def normalize_school_name(value, field='Name'):
    raw=' '.join(str(value or '').strip().split()).upper()
    return re.sub(r"[^A-Z .'\-]", '', raw)[:160]

ROMAN_TO_INT = {'I':1,'II':2,'III':3,'IV':4,'V':5,'VI':6,'VII':7,'VIII':8,'IX':9,'X':10,'XI':11,'XII':12}
INT_TO_ROMAN = {v:k for k,v in ROMAN_TO_INT.items()}

def class_number(value):
    raw=''.join(str(value or '').strip().upper().split())
    if raw.isdigit():
        n=int(raw)
    else:
        n=ROMAN_TO_INT.get(raw, 0)
    return n

def normalize_class(value):
    n=class_number(value)
    if 1 <= n <= 10:
        return INT_TO_ROMAN[n]
    raw=' '.join(str(value or '').strip().split()).upper()
    return re.sub(r'[^A-Z0-9 \-]', '', raw)[:10]

def normalize_section(value):
    raw=' '.join(str(value or '').strip().split()).upper()
    return re.sub(r'[^A-Z0-9]', '', raw)[:10]

def language_label(code):
    return {'telugu':'Telugu','hindi':'Hindi','sanskrit':'Sanskrit'}.get(str(code or '').lower(), str(code or '').title())

def subject_options_for_class(class_name, second_language=None, third_language=None):
    n=class_number(class_name)
    if 5<=n<=8:
        return [('eng','English (1st Language)'),
                (f'lang2_{str(second_language).lower()}','2nd Language: '+language_label(second_language)) if second_language else ('lang2_placeholder','2nd Language'),
                (f'lang3_{str(third_language).lower()}','3rd Language: '+language_label(third_language)) if third_language else ('lang3_placeholder','3rd Language'),
                ('computers','Computers'),('math','Mathematics'),('social','Social Science'),('science','Science')]
    if 9<=n<=10:
        return [('math','Mathematics'),('chemistry','Chemistry'),('biology','Biology'),('physics','Physics'),('social','Social Science'),('eng','English (1st Language)'),
                (f'lang2_{str(second_language).lower()}','2nd Language: '+language_label(second_language)) if second_language else ('lang2_placeholder','2nd Language'),('it','Information Technology')]
    return []

def class_list():
    return ['I','II','III','IV','V','VI','VII','VIII','IX','X','XI','XII']

def seed_data():
    if not db.session.query(Account).filter_by(role='admin').first():
        u=os.getenv('INITIAL_ADMIN_USERNAME','admin'); p=os.getenv('INITIAL_ADMIN_PASSWORD','ChangeThisImmediately123!')
        db.session.add(Account(username=u,password_hash=generate_password_hash(p),role='admin',display_name='SCHOOL ADMINISTRATOR',must_change_password=True))
    if not Subject.query.first():
        subjects=[]
        base_common=[('eng','English'),('math','Mathematics')]
        for code,name in base_common: subjects.append(Subject(code=code,name=name,class_band='5-10'))
        for code,name in [('social','Social Science'),('science','Science'),('computers','Computers'),('it','Information Technology'),('physics','Physics'),('chemistry','Chemistry'),('biology','Biology')]: subjects.append(Subject(code=code,name=name,class_band='5-10'))
        for lang in ['telugu','hindi','sanskrit']:
            subjects.append(Subject(code=f'lang2_{lang}',name=lang.title(),class_band='5-10',language_group='second'))
            subjects.append(Subject(code=f'lang3_{lang}',name=lang.title(),class_band='5-8',language_group='third'))
        db.session.add_all(subjects)
    # Normalize language subject names so identical languages remain visibly distinct
    # between 2nd and 3rd language slots.
    for code, label in [
        ('lang2_telugu','Telugu (2nd Language)'),
        ('lang2_hindi','Hindi (2nd Language)'),
        ('lang2_sanskrit','Sanskrit (2nd Language)'),
        ('lang3_telugu','Telugu (3rd Language)'),
        ('lang3_hindi','Hindi (3rd Language)'),
        ('lang3_sanskrit','Sanskrit (3rd Language)')
    ]:
        subj=Subject.query.filter_by(code=code).first()
        if subj: subj.name=label
    if not Exam.query.first():
        db.session.add_all([Exam(name='PT-1',max_marks=40,order_index=1),Exam(name='PT-2',max_marks=80,order_index=2),Exam(name='PT-3',max_marks=40,order_index=3),Exam(name='Final Examination',max_marks=80,order_index=4,is_final=True)])
    if SchoolSetting.query.filter_by(key='academic_session').first() is None:
        db.session.add(SchoolSetting(key='academic_session',value=os.getenv('ACADEMIC_SESSION','2026-27')))
    # Backfill missing accounts for legacy students so every student record has a login.
    legacy_password=os.getenv('LEGACY_STUDENT_DEFAULT_PASSWORD','ChangeThisImmediately123!')
    for st in Student.query.filter(Student.account_id.is_(None), Student.active.is_(True)).limit(5000).all():
        base=''.join(ch.lower() if ch.isalnum() else '_' for ch in (st.admission_number or f'student_{st.id}')).strip('_') or f'student_{st.id}'
        username=base
        n=1
        while Account.query.filter_by(username=username).first():
            n+=1; username=f'{base}_{n}'
        acct=Account(username=username,password_hash=generate_password_hash(legacy_password),role='student',display_name=st.name,must_change_password=True,active=st.active)
        db.session.add(acct); db.session.flush(); st.account_id=acct.id
    for t in Teacher.query.filter(Teacher.account_id.is_(None), Teacher.active.is_(True)).limit(1000).all():
        base=''.join(ch.lower() if ch.isalnum() else '_' for ch in (t.username or t.name or f'teacher_{t.id}')).strip('_') or f'teacher_{t.id}'
        username=base; n=1
        while Account.query.filter_by(username=username).first():
            n+=1; username=f'{base}_{n}'
        acct=Account(username=username,password_hash=generate_password_hash(legacy_password),role='teacher',display_name=t.name,must_change_password=True,active=t.active)
        db.session.add(acct); db.session.flush(); t.account_id=acct.id
    db.session.commit()



def ensure_announcement_schema():
    """Ensure targeted announcement column/index exists on older databases."""
    try:
        db.session.execute(text("ALTER TABLE announcement ADD COLUMN IF NOT EXISTS parent_id INTEGER REFERENCES parent(id) ON DELETE CASCADE"))
        db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_announcement_parent ON announcement(parent_id, published)"))
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise


def ensure_fee_schema():
    """Create the lightweight fee tables on-demand if migration 007 was not run."""
    try:
        db.metadata.create_all(bind=db.engine, tables=[FeeStructure.__table__, FeeStructureDocument.__table__, FeePaymentWindow.__table__])
        db.session.execute(text('ALTER TABLE fee_invoice ADD COLUMN IF NOT EXISTS academic_session VARCHAR(30)'))
        db.session.execute(text('ALTER TABLE announcement ADD COLUMN IF NOT EXISTS parent_id INTEGER REFERENCES parent(id) ON DELETE CASCADE'))
        db.session.execute(text('CREATE INDEX IF NOT EXISTS ix_announcement_parent ON announcement(parent_id, published)'))
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise


with app.app_context():
    ensure_announcement_schema()
    # Production/Render uses the already-migrated Supabase schema.  Auto-DDL is
    # opt-in for local development only, which avoids repeated schema checks on
    # every Gunicorn worker boot.
    if os.getenv('AUTO_CREATE_SCHEMA','false').lower() == 'true':
        db.create_all()
    if os.getenv('AUTO_SEED','true').lower() == 'true':
        seed_data()



def grade_for_percent(pct):
    try: pct=float(pct)
    except Exception: return 'E'
    if pct >= 91: return 'A1'
    if pct >= 81: return 'A2'
    if pct >= 71: return 'B1'
    if pct >= 61: return 'B2'
    if pct >= 51: return 'C1'
    if pct >= 41: return 'C2'
    if pct >= 33: return 'D'
    return 'E'

def pass_fail_for_percent(pct):
    return 'PASS' if float(pct) >= 33 else 'REQUIRES ATTENTION'

def subjects_for_class(class_name, second_language=None, third_language=None):
    codes=[c for c,_ in subject_options_for_class(class_name,second_language,third_language) if not c.endswith('_placeholder')]
    return [Subject.query.filter_by(code=c).first() for c in codes if Subject.query.filter_by(code=c).first()]

def working_days_between(start,end):
    rows=SchoolCalendar.query.filter(SchoolCalendar.date>=start,SchoolCalendar.date<=end).all()
    overrides={r.date:r.is_working for r in rows}
    cur=start; count=0
    while cur<=end:
        count += 1 if overrides.get(cur,cur.weekday()<5) else 0
        cur += timedelta(days=1)
    return count

def attendance_summary(student_id, start, end):
    wd=working_days_between(start,end)
    ats=Attendance.query.filter_by(student_id=student_id).filter(Attendance.date>=start,Attendance.date<=end).all()
    present=sum(1 for a in ats if a.status in {'present','late'})
    absent=sum(1 for a in ats if a.status=='absent')
    pct=round((present/wd)*100,2) if wd else 0
    return {'working_days':wd,'present':present,'absent':absent,'percentage':pct}

def attendance_percentage_map(student_ids, start, end):
    ids=[int(x) for x in student_ids if x is not None]
    if not ids:
        return {}
    wd=working_days_between(start,end)
    if wd <= 0:
        return {sid:0 for sid in ids}
    rows=(db.session.query(Attendance.student_id, func.count(Attendance.id))
          .filter(Attendance.student_id.in_(ids), Attendance.date>=start, Attendance.date<=end, Attendance.status.in_(['present','late']))
          .group_by(Attendance.student_id).all())
    return {sid: round((count/wd)*100,2) for sid,count in rows}

def school_timezone():
    try:
        return ZoneInfo(app.config.get('APP_TIMEZONE','Asia/Kolkata'))
    except Exception:
        return ZoneInfo('Asia/Kolkata')

def _configured_clock_time():
    row=db.session.get(SchoolClock,1)
    raw=(row.override_time or '').strip() if row else ''
    if not raw:
        return None
    for fmt in ('%H:%M','%H:%M:%S'):
        try:
            return datetime.strptime(raw,fmt).time()
        except ValueError:
            pass
    return None

def school_now():
    now=datetime.now(school_timezone())
    override=_configured_clock_time()
    if override is not None:
        return datetime.combine(now.date(),override,tzinfo=school_timezone())
    return now

def school_time(): return school_now().time().replace(microsecond=0)

def school_date(): return school_now().date()

def attendance_status_for_time(value=None):
    t=value or school_time()
    def parse_env(name, default):
        raw=os.getenv(name,default)
        try: return datetime.strptime(raw,'%H:%M').time()
        except ValueError: return datetime.strptime(default,'%H:%M').time()
    present_from=parse_env('ATTENDANCE_PRESENT_FROM','07:30')
    late_after=parse_env('ATTENDANCE_LATE_AFTER','08:30')
    absent_after=parse_env('ATTENDANCE_ABSENT_AFTER','09:00')
    if t < present_from: return 'present'
    if t < late_after: return 'present'
    if t < absent_after: return 'late'
    return 'absent'

def school_year_bounds(): return date(school_date().year if school_date().month>=4 else school_date().year-1,4,1), date(school_date().year+1 if school_date().month>=4 else school_date().year,3,31)

def allowed_students_for_account(acct):
    if acct.role=='admin': return Student.query.filter_by(active=True)
    if acct.role=='teacher':
        t=Teacher.query.filter_by(account_id=acct.id).first()
        if not t: return Student.query.filter(False)
        assignments=TeacherAssignment.query.filter_by(teacher_id=t.id).all()
        from sqlalchemy import or_, and_
        clauses=[and_(Student.class_name==a.class_name,Student.section==a.section) for a in assignments]
        return Student.query.filter(or_(*clauses)) if clauses else Student.query.filter(False)
    if acct.role=='student':
        s=Student.query.filter_by(account_id=acct.id).first(); return Student.query.filter_by(id=s.id) if s else Student.query.filter(False)
    if acct.role=='parent':
        p=Parent.query.filter_by(account_id=acct.id).first(); ids=[x.student_id for x in ParentStudent.query.filter_by(parent_id=p.id).all()] if p else []
        return Student.query.filter(Student.id.in_(ids)) if ids else Student.query.filter(False)
    return Student.query.filter(False)

@app.get('/healthz')
def healthz():
    try:
        db.session.execute(text('SELECT 1'))
        return jsonify({'status':'ok','database':'ok'})
    except Exception:
        db.session.rollback()
        return jsonify({'status':'degraded','database':'error'}),503


def _local_ai_fallback(question, acct, students, attendance_pct, academic_pct):
    q=question.lower().strip()
    # Deterministic, scope-safe school assistant. It never depends on an API quota.
    if 'how many students' in q and ('attendance' in q or 'absent' in q):
        m=re.search(r'(?:less than|below|under)\s+(\d+(?:\.\d+)?)',q)
        if m:
            t=float(m.group(1)); n=sum(1 for s in students if attendance_pct.get(s.id,0)<t)
            return f'{n} student{"s" if n!=1 else ""} in your allowed scope have attendance below {t:g}%. '
    if ('which students' in q or 'who' in q or 'list' in q) and ('attendance' in q or 'absent' in q):
        m=re.search(r'(?:less than|below|under)\s+(\d+(?:\.\d+)?)',q)
        if m:
            t=float(m.group(1)); hits=[s for s in students if attendance_pct.get(s.id,0)<t]
            return f'\n'.join([f'{len(hits)} students match your attendance filter.']+[f'{s.name} ({s.class_name}{("-"+s.section) if s.section else ""}) — {attendance_pct.get(s.id,0):g}%' for s in hits[:100]]) if hits else 'No students match your attendance filter.'
    if ('which students' in q or 'who' in q or 'list' in q) and ('academic' in q or 'academics' in q or 'marks' in q or 'results' in q):
        m=re.search(r'(?:less than|below|under)\s+(\d+(?:\.\d+)?)',q)
        if m:
            t=float(m.group(1)); hits=[s for s in students if academic_pct.get(s.id,0)<t]
            return '\n'.join([f'{len(hits)} students match your academic filter.']+[f'{s.name} ({s.class_name}{("-"+s.section) if s.section else ""}) — {academic_pct.get(s.id,0):g}%' for s in hits[:100]]) if hits else 'No students match your academic filter.'
    if 'how many students' in q and ('academic' in q or 'academics' in q or 'marks' in q or 'results' in q):
        m=re.search(r'(?:less than|below|under)\s+(\d+(?:\.\d+)?)',q)
        if m:
            t=float(m.group(1)); n=sum(1 for s in students if academic_pct.get(s.id,0)<t)
            return f'{n} student{"s" if n!=1 else ""} in your allowed scope have academic performance below {t:g}%. '
    if 'how many students' in q:
        return f'You can access {len(students)} student records in your current scope.'
    return 'I can answer questions using the school data available to your account. Try asking about attendance, academics/results, fees, students, classes, or parents.'


@app.get('/')
def index():
    if session.get('account_id'): return redirect(url_for('dashboard'))
    return render_template('landing.html')

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method=='POST':
        username=request.form.get('username','').strip(); password=request.form.get('password','')
        acct=Account.query.filter(func.lower(Account.username)==username.lower()).first()
        if not acct or not acct.active or not check_password_hash(acct.password_hash,password):
            return render_template('login.html',error='Invalid username or password.'),401
        session.clear(); session['account_id']=acct.id; session['role']=acct.role; session['csrf']=secrets.token_urlsafe(24); acct.last_login=datetime.utcnow(); log_audit('login', 'Account', acct.id); db.session.commit()
        return redirect(request.args.get('next') or url_for('dashboard'))
    return render_template('login.html')

@app.get('/logout')
def logout(): session.clear(); return redirect(url_for('index'))

PROFILE_DIR = BASE_DIR / 'static' / 'uploads' / 'profiles'
PROFILE_DIR.mkdir(parents=True, exist_ok=True)

def profile_picture_url(account):
    if not account or not account.profile_picture_path:
        return None
    return url_for('static', filename=account.profile_picture_path)

@app.post('/account/profile-picture')
@login_required()
def upload_profile_picture():
    acct=current_account()
    upload=request.files.get('profile_picture')
    if not upload or not upload.filename:
        flash('CHOOSE A PROFILE PICTURE FIRST.','error')
        return redirect(url_for('change_credentials'))
    raw=upload.read()
    if len(raw) > 3 * 1024 * 1024:
        flash('PROFILE PICTURE MUST BE 3 MB OR SMALLER.','error')
        return redirect(url_for('change_credentials'))
    if np is None or cv2 is None:
        flash('PROFILE PICTURES REQUIRE NUMPY AND OPENCV.','error')
        return redirect(url_for('change_credentials'))
    arr=np.frombuffer(raw,dtype=np.uint8)
    img=cv2.imdecode(arr,cv2.IMREAD_COLOR)
    if img is None:
        flash('INVALID IMAGE. USE JPG, PNG OR WEBP.','error')
        return redirect(url_for('change_credentials'))
    h,w=img.shape[:2]
    if h < 64 or w < 64:
        flash('PROFILE PICTURE IS TOO SMALL.','error')
        return redirect(url_for('change_credentials'))
    side=min(h,w)
    y=(h-side)//2; x=(w-side)//2
    img=img[y:y+side,x:x+side]
    img=cv2.resize(img,(512,512),interpolation=cv2.INTER_AREA)
    ok,encoded=cv2.imencode('.jpg',img,[int(cv2.IMWRITE_JPEG_QUALITY),90])
    if not ok:
        flash('COULD NOT PROCESS THAT IMAGE.','error')
        return redirect(url_for('change_credentials'))
    filename=f"account-{acct.id}-{secrets.token_hex(8)}.jpg"
    target=PROFILE_DIR/filename
    target.write_bytes(encoded.tobytes())
    if acct.profile_picture_path:
        old=BASE_DIR/'static'/acct.profile_picture_path.replace('static/','',1) if acct.profile_picture_path.startswith('static/') else BASE_DIR/'static'/acct.profile_picture_path
        try:
            if old.exists() and old != target: old.unlink()
        except OSError:
            pass
    acct.profile_picture_path=f"uploads/profiles/{filename}"
    log_audit('profile_picture_updated','Account',acct.id,{'filename':filename})
    db.session.commit()
    flash('PROFILE PICTURE UPDATED.','success')
    return redirect(url_for('change_credentials'))

@app.post('/account/profile-picture/remove')
@login_required()
def remove_profile_picture():
    acct=current_account()
    if acct.profile_picture_path:
        old=BASE_DIR/'static'/acct.profile_picture_path
        try:
            if old.exists(): old.unlink()
        except OSError:
            pass
        acct.profile_picture_path=None
        log_audit('profile_picture_removed','Account',acct.id)
        db.session.commit()
    flash('PROFILE PICTURE REMOVED.','success')
    return redirect(url_for('change_credentials'))

@app.route('/account', methods=['GET','POST'])
@login_required()
def change_credentials():
    acct=current_account()
    if request.method=='POST':
        current=request.form.get('current_password',''); new_user=request.form.get('username','').strip(); new_password=request.form.get('new_password',''); confirm=request.form.get('confirm_password','')
        if not new_user: return render_template('account.html',error='Username is required.',account=acct,first_time=acct.must_change_password)
        if not check_password_hash(acct.password_hash,current): return render_template('account.html',error='Current password is incorrect.',account=acct,first_time=acct.must_change_password)
        if len(new_password)<8 or new_password!=confirm: return render_template('account.html',error='Use an 8+ character password and make sure both new passwords match.',account=acct,first_time=acct.must_change_password)
        if new_user!=acct.username and Account.query.filter(func.lower(Account.username)==new_user.lower()).first(): return render_template('account.html',error='Username already exists.',account=acct,first_time=acct.must_change_password)
        acct.username=new_user; acct.password_hash=generate_password_hash(new_password); acct.must_change_password=False; log_audit('credentials_changed','Account',acct.id); db.session.commit(); return redirect(url_for('dashboard'))
    return render_template('account.html',account=acct,first_time=acct.must_change_password)

@app.get('/dashboard')
@login_required()
def dashboard():
    acct=current_account(); q=allowed_students_for_account(acct)
    student_count=q.count()
    start,end=school_year_bounds(); today=min(school_date(),end)
    students=q.order_by(Student.class_name,Student.section,Student.roll_number,Student.name).limit(8).all()
    attendance_map = attendance_percentage_map([st.id for st in students], start, today)
    data=[{'name':st.name,'class_section':f'{st.class_name}-{st.section}' if st.section else st.class_name,
           'attendance':attendance_map.get(st.id,0)} for st in students]
    year_total=round(sum(x['attendance'] for x in data)/len(data),2) if data else 0

    audience_map={'admin':'admin','teacher':'teachers','student':'students','parent':'parents'}
    aq=Announcement.query.filter(Announcement.published.is_(True))
    if acct.role == 'admin':
        aq=aq.filter(Announcement.audience.in_(['all','admin']))
    else:
        aq=aq.filter(Announcement.audience.in_(['all',audience_map.get(acct.role)]))
    announcements=aq.order_by(Announcement.published_at.desc()).limit(5).all()

    class_teacher='—'
    if acct.role=='student':
        st=Student.query.filter_by(account_id=acct.id).first()
        if st:
            ta=TeacherAssignment.query.filter_by(class_name=st.class_name,section=st.section).first()
            if ta:
                t=db.session.get(Teacher,ta.teacher_id); class_teacher=t.name if t else '—'
    children_count=student_count if acct.role == 'parent' else None
    return render_template('dashboard.html',account=acct,role=acct.role,student_count=student_count,
        teacher_count=Teacher.query.filter_by(active=True).count() if acct.role=='admin' else None,
        attendance=data,year_attendance=year_total,announcements=announcements,
        working_today=is_working_day(school_date()),class_teacher=class_teacher,children_count=children_count)

@app.get('/admin')
@admin_required
def admin_home(): return redirect(url_for('dashboard'))

@app.get('/admin/register')
@admin_required
def admin_register_page():
    return render_template('register_student.html', class_options=class_list(), form={})

@app.post('/admin/register')
@admin_required
def admin_register_save():
    d=request.form; name=normalize_school_name(d.get('name')); adm=str(d.get('admission_number','')).strip().upper(); roll=str(d.get('roll_number','')).strip().upper()
    cls=normalize_class(d.get('class_name')); sec=normalize_section(d.get('section')); user=str(d.get('username','')).strip(); password=d.get('password','')
    second=(d.get('second_language') or '').strip().lower() or None; third=(d.get('third_language') or '').strip().lower() or None
    if class_number(cls)>=9: third=None
    section_required=class_number(cls) not in {11,12}
    if not all([name,adm,roll,cls,user]) or (section_required and not sec) or len(password)<8:
        flash('NAME, ADMISSION NUMBER, ROLL NUMBER, CLASS, USERNAME AND AN 8+ CHARACTER TEMPORARY PASSWORD ARE REQUIRED. SECTION IS REQUIRED EXCEPT FOR XI/XII.','error')
        return render_template('register_student.html',class_options=class_list(),form=d)
    if class_number(cls) in range(5,9) and (not second or not third):
        flash('2ND AND 3RD LANGUAGE ARE REQUIRED FOR CLASSES V–VIII.','error'); return render_template('register_student.html',class_options=class_list(),form=d)
    if class_number(cls) in range(9,11) and not second:
        flash('2ND LANGUAGE IS REQUIRED FOR CLASSES IX–X.','error'); return render_template('register_student.html',class_options=class_list(),form=d)
    if Student.query.filter_by(admission_number=adm).first():
        flash('ADMISSION NUMBER ALREADY EXISTS.','error'); return render_template('register_student.html',class_options=class_list(),form=d)
    if Account.query.filter(func.lower(Account.username)==user.lower()).first():
        flash('USERNAME ALREADY EXISTS.','error'); return render_template('register_student.html',class_options=class_list(),form=d)
    session['pending_student_registration']={'name':name,'admission_number':adm,'roll_number':roll,'class_name':cls,'section':sec,'username':user,'password_hash':generate_password_hash(password),'second_language':second,'third_language':third}
    if not face_available():
        acct=Account(username=user,password_hash=generate_password_hash(password),role='student',display_name=name,must_change_password=True)
        db.session.add(acct); db.session.flush()
        st=Student(name=name,admission_number=adm,roll_number=roll,class_name=cls,section=sec,second_language=second,third_language=third,account_id=acct.id)
        db.session.add(st); log_audit('student_created','Student',st.id,{'class':cls,'section':sec,'username':user,'face_frames':0}); db.session.commit(); session.pop('pending_student_registration',None)
        flash(f'STUDENT CREATED. USERNAME: {acct.username}. FACE ID CAN BE ADDED LATER.','success'); return redirect(url_for('admin_students'))
    return redirect(url_for('register_student_face'))

@app.get('/admin/register/face')
@admin_required
def register_student_face():
    pending=session.get('pending_student_registration')
    if not pending:
        flash('START STUDENT REGISTRATION FIRST.','error')
        return redirect(url_for('admin_register_page'))
    return render_template('register_face_capture.html', pending=pending)

@app.post('/admin/register/face/complete')
@admin_required
def register_student_face_complete():
    pending=session.get('pending_student_registration')
    if not pending:
        return jsonify({'error':'Registration session expired. Start again.'}),403
    frames=request.form.getlist('frames')
    if len(frames)<8:
        return jsonify({'error':'FACE ID IS REQUIRED. CAPTURE AT LEAST 8 VALID FRAMES.'}),400
    encs=[]
    for raw in frames:
        try:
            encs.append(encode_frame(raw))
        except Exception:
            continue
    if len(encs)<8:
        return jsonify({'error':f'Only {len(encs)} valid face frames were detected. Capture at least 8 good frames.'}),400
    if Student.query.filter_by(admission_number=pending['admission_number']).first():
        session.pop('pending_student_registration',None); return jsonify({'error':'Admission number already exists.'}),409
    if Account.query.filter(func.lower(Account.username)==pending['username'].lower()).first():
        session.pop('pending_student_registration',None); return jsonify({'error':'Username already exists.'}),409
    acct=Account(username=pending['username'],password_hash=pending['password_hash'],role='student',display_name=pending['name'],must_change_password=True)
    db.session.add(acct); db.session.flush()
    s=Student(name=pending['name'],admission_number=pending['admission_number'],roll_number=pending['roll_number'],class_name=pending['class_name'],section=pending['section'],second_language=pending['second_language'],third_language=pending['third_language'],account_id=acct.id,face_encoding_json=json.dumps(encs),face_trained=True)
    db.session.add(s); db.session.flush()
    log_audit('student_created','Student',s.id,{'class':s.class_name,'section':s.section,'username':acct.username,'face_frames':len(encs)})
    db.session.commit(); session.pop('pending_student_registration',None)
    flash(f'STUDENT CREATED WITH FACE ID. USERNAME: {acct.username}. FIRST LOGIN REQUIRES A CREDENTIAL CHANGE.','success')
    return jsonify({'ok':True,'redirect':url_for('admin_students')})

@app.get('/admin/students')
@admin_required
def admin_students():
    raw_day=request.args.get('date') or school_date().isoformat()
    try: day=datetime.strptime(raw_day,'%Y-%m-%d').date()
    except ValueError: day=school_date()
    cls=normalize_class(request.args.get('class_name','')); sec=normalize_section(request.args.get('section','')); search=request.args.get('q','').strip()
    q=Student.query.filter_by(active=True)
    if cls: q=q.filter(Student.class_name==cls)
    if sec: q=q.filter(Student.section==sec)
    if search:
        like=f'%{search}%'; q=q.filter(or_(Student.name.ilike(like),Student.admission_number.ilike(like),Student.roll_number.ilike(like)))
    total=q.count(); page=max(1,request.args.get('page',1,type=int)); per_page=50
    students=q.order_by(Student.class_name,Student.section,Student.roll_number,Student.name).offset((page-1)*per_page).limit(per_page).all()
    ids=[x.id for x in students]; ats=Attendance.query.filter(Attendance.date==day,Attendance.student_id.in_(ids)).all() if ids else []
    att_by={a.student_id:a for a in ats}; rows=[{'student':st,'attendance':att_by.get(st.id)} for st in students]
    classes=[x for x in class_list() if Student.query.filter_by(active=True,class_name=x).first()]
    sections=sorted({x.section for x in students if x.section})
    return render_template('students.html',rows=rows,day=day,class_name=cls,section=sec,classes=classes,sections=sections,page=page,per_page=per_page,total=total,search=search)

@app.route('/admin/teachers', methods=['GET','POST'])
@admin_required
def admin_teachers():
    if request.method=='POST':
        name=normalize_school_name(request.form.get('name'))
        username=request.form.get('username','').strip()
        password=request.form.get('password','')
        phone=request.form.get('phone','').strip()
        if not name or not username or len(password)<8:
            flash('NAME, USERNAME AND AN 8+ CHARACTER TEMPORARY PASSWORD ARE REQUIRED.','error')
        elif Account.query.filter(func.lower(Account.username)==username.lower()).first():
            flash('USERNAME ALREADY EXISTS.','error')
        else:
            acct=Account(username=username,password_hash=generate_password_hash(password),role='teacher',display_name=name,must_change_password=True)
            db.session.add(acct); db.session.flush()
            t=Teacher(name=name,phone=phone,account_id=acct.id); db.session.add(t); db.session.flush()
            log_audit('teacher_created','Teacher',t.id,{'username':username}); db.session.commit(); flash('TEACHER ACCOUNT CREATED.','success')
    teachers=Teacher.query.order_by(Teacher.name).all()
    assignment_counts={t.id:TeacherAssignment.query.filter_by(teacher_id=t.id).count() for t in teachers}
    return render_template('teachers.html',teachers=teachers,assignment_counts=assignment_counts)

@app.route('/admin/announcements', methods=['GET','POST'])
@admin_required
def announcements():
    if request.method=='POST':
        title=request.form.get('title','').strip(); message=request.form.get('message','').strip(); audience=request.form.get('audience','all').strip(); publish=request.form.get('publish')=='1'
        allowed={'public','all','admin','students','teachers','parents'}
        if not title or not message or audience not in allowed:
            flash('Title, message and one valid audience are required.','error')
        else:
            a=Announcement(title=title,message=message,audience=audience,created_by=current_account().username,published=publish)
            if publish: a.published_at=datetime.utcnow()
            db.session.add(a); db.session.flush(); log_audit('announcement_created','Announcement',a.id,{'audience':audience,'published':publish})
            db.session.commit(); flash('Announcement created for the selected audience.','success')
    return render_template('announcements.html',announcements=Announcement.query.order_by(Announcement.created_at.desc()).limit(200).all())

@app.post('/admin/announcements/<int:aid>/publish')
@admin_required
def publish_announcement(aid):
    a=db.session.get(Announcement,aid) or abort(404); a.published=True; a.published_at=datetime.utcnow(); log_audit('announcement_published','Announcement',aid); db.session.commit(); return redirect(url_for('announcements'))

@app.post('/admin/announcements/<int:aid>/delete')
@admin_required
def delete_announcement(aid):
    a=db.session.get(Announcement,aid) or abort(404)
    db.session.delete(a); log_audit('announcement_deleted','Announcement',aid,{'title':a.title}); db.session.commit()
    flash('Announcement deleted.','success')
    return redirect(url_for('announcements'))

@app.get('/teacher')
@staff_required
def teacher_home(): return redirect(url_for('dashboard'))

@app.get('/students')
@login_required()
def students_self():
    acct=current_account(); rows=allowed_students_for_account(acct).all(); return render_template('portal_students.html',students=rows,role=acct.role)

@app.get('/api/school-day')
@staff_required
def school_day_api():
    try:
        today=school_date()
        override=SchoolCalendar.query.filter_by(date=today).first()
        effective=school_time()
        return jsonify({'date':today.isoformat(),'is_working':is_working_day(today),'reason':override.reason if override else None,'override':bool(override),'time':effective.strftime('%H:%M'),'live_server_time':now_local().strftime('%H:%M'),'using_override':bool(get_school_clock_override()),'attendance_from':os.getenv('ATTENDANCE_PRESENT_FROM','07:30'),'late_after':os.getenv('ATTENDANCE_LATE_AFTER','08:30'),'absent_after':os.getenv('ATTENDANCE_ABSENT_AFTER','09:00')})
    except Exception:
        db.session.rollback()
        today=school_date(); return jsonify({'date':today.isoformat(),'is_working':weekly_default_is_working(today),'reason':None,'override':False,'time':school_time().strftime('%H:%M'),'live_server_time':now_local().strftime('%H:%M'),'using_override':False,'calendar_fallback':True,'attendance_from':os.getenv('ATTENDANCE_PRESENT_FROM','07:30'),'late_after':os.getenv('ATTENDANCE_LATE_AFTER','08:30'),'absent_after':os.getenv('ATTENDANCE_ABSENT_AFTER','09:00')})

@app.get('/attendance/scan')
@staff_required
def attendance_scan_page():
    if not ip_allowed('SCANNER_ALLOWED_IPS'):
        abort(403)
    return render_template('attendance_scan.html',face_available=face_available(),scanner_ip_restricted=bool(os.getenv('SCANNER_ALLOWED_IPS','').strip()))

@app.get('/attendance')
@staff_required
def attendance_page():
    try: day=datetime.strptime(request.args.get('date',school_date().isoformat()),'%Y-%m-%d').date()
    except ValueError: day=school_date()
    q=allowed_students_for_account(current_account()); cls=normalize_class(request.args.get('class_name','')); sec=normalize_section(request.args.get('section',''))
    if cls: q=q.filter(Student.class_name==cls)
    if sec: q=q.filter(Student.section==sec)
    search=request.args.get('q','').strip()
    if search:
        like=f'%{search}%'; q=q.filter(or_(Student.name.ilike(like),Student.admission_number.ilike(like)))
    total=q.count(); page=max(1,request.args.get('page',1,type=int)); per_page=100
    students=q.order_by(Student.class_name,Student.section,Student.roll_number,Student.name).offset((page-1)*per_page).limit(per_page).all()
    ids=[x.id for x in students]; ats=Attendance.query.filter(Attendance.date==day,Attendance.student_id.in_(ids)).all() if ids else []
    att_by={a.student_id:a for a in ats}; rows=[{'student':st,'attendance':att_by.get(st.id)} for st in students]
    classes=sorted({x.class_name for x in allowed_students_for_account(current_account()).with_entities(Student.class_name).distinct().all()},key=lambda x:(class_number(x) or 99,x))
    return render_template('attendance.html',rows=rows,day=day,face_available=face_available(),classes=classes,class_name=cls,section=sec,page=page,total=total,per_page=per_page,search=search)

@app.post('/api/attendance/mark')
@staff_required
def mark_attendance():
    d=request.get_json(silent=True) or {}; status=str(d.get('status','present')).lower()
    if status not in {'present','late','absent'}: return jsonify({'error':'Status must be present, late or absent.'}),400
    try: sid=int(d.get('student_id')); day=datetime.strptime(d.get('date',school_date().isoformat()),'%Y-%m-%d').date()
    except (TypeError,ValueError): return jsonify({'error':'Student and date are required.'}),400
    s=Student.query.filter_by(id=sid,active=True).first(); acct=current_account()
    if not s: abort(404)
    if acct.role!='admin' and allowed_students_for_account(acct).filter_by(id=sid).first() is None: abort(403)
    a=Attendance.query.filter_by(student_id=sid,date=day).first() or Attendance(student_id=sid,date=day)
    a.status=status; a.time_in=school_time() if status in {'present','late'} else None; a.source='manual'; a.marked_by=acct.username; a.note=(d.get('note') or '').strip()[:500]; db.session.add(a)
    log_audit('attendance_marked','Student',sid,{'status':status,'date':str(day),'source':'manual'}); db.session.commit(); return jsonify({'ok':True,'status':status})

@app.post('/api/attendance/recognize')
@staff_required
def recognize():
    if not ip_allowed('SCANNER_ALLOWED_IPS'):
        return jsonify({'error':'Scanner access is restricted to the configured scanner network.'}),403
    if not face_available():
        return jsonify({'error':'Face recognition dependencies are not available on this server.'}),503
    try:
        data=request.get_json() or {}
        enc=encode_frame(data.get('image'))
    except Exception as exc:
        return jsonify({'error':str(exc) or 'Invalid camera frame.'}),400

    # Compare against every eligible student and choose the single closest match.
    # The previous scanner stopped at the first student within tolerance, which could
    # mark the wrong student when multiple face encodings were close enough.
    candidates=[]
    students=allowed_students_for_account(current_account()).filter(Student.face_trained.is_(True)).all()
    for s in students:
        if not s.face_encoding_json:
            continue
        try:
            distance=match_distance(enc,json.loads(s.face_encoding_json))
        except Exception:
            distance=None
        if distance is not None:
            candidates.append((distance,s))

    if not candidates:
        return jsonify({'matched':False})

    distance, s=min(candidates,key=lambda item:item[0])
    if distance > 0.48:
        return jsonify({'matched':False})

    day=school_date()
    a=Attendance.query.filter_by(student_id=s.id,date=day).first()
    if a:
        # A second scan must never downgrade a valid earlier scan from Present/Late
        # to a later status such as Absent. Attendance is established by the first
        # successful scan for that school day; staff can correct it manually.
        return jsonify({'matched':True,'student':s.name,'status':a.status,'already_marked':True})

    status=attendance_status_for_time()
    a=Attendance(student_id=s.id,date=day)
    a.status=status
    a.source='face'
    a.marked_by=current_account().username
    a.time_in=school_time() if status in {'present','late'} else None
    db.session.add(a)
    log_audit('face_attendance','Student',s.id,{'status':status,'distance':round(distance,4)})
    db.session.commit()
    return jsonify({'matched':True,'student':s.name,'status':status,'already_marked':False})

@app.route('/admin/students/<int:sid>/face', methods=['GET','POST'])
@admin_required
def student_face(sid):
    s=db.session.get(Student,sid) or abort(404)
    if request.method=='POST':
        frames=request.form.getlist('frames'); encs=[]
        for f in frames:
            try: encs.append(encode_frame(f))
            except Exception: pass
        if not encs: return render_template('face_capture.html',student=s,error='No valid face frames received.')
        s.face_encoding_json=json.dumps(encs); s.face_trained=True; log_audit('face_model_updated','Student',sid,{'frames':len(encs)}); db.session.commit(); flash('Face model updated.','success'); return redirect(url_for('admin_students'))
    return render_template('face_capture.html',student=s)

@app.get('/academics')
@staff_required
def academics():
    acct=current_account(); allowed=allowed_students_for_account(acct)
    subject_param=request.args.get('subject','').strip()
    # Teachers are class teachers only: a class teacher can enter marks for
    # every subject of the students in their assigned class/section.
    # Subject-level teacher assignment is intentionally not used.
    subjects_q=Subject.query.order_by(Subject.name)
    subjects=subjects_q.all()
    if subject_param and any(x.code==subject_param for x in subjects):
        subjects=[x for x in subjects if x.code==subject_param]
    students=allowed.order_by(Student.class_name,Student.section,Student.roll_number,Student.name).limit(500).all()
    exams=Exam.query.order_by(Exam.order_index).all()
    return render_template('academics.html',exams=exams,subjects=subjects,all_subjects=subjects_q.all() if subject_param else subjects,students=students,selected_subject=subject_param)

@app.post('/api/marks')
@staff_required
def save_mark():
    d=request.get_json() or {}; sid=int(d['student_id']); exam=db.session.get(Exam,int(d['exam_id'])); subj=Subject.query.filter_by(code=d['subject_code']).first(); val=float(d['marks']) if d.get('marks') not in ('',None) else None
    student=allowed_students_for_account(current_account()).filter_by(id=sid).first()
    if not student or not exam or not subj: abort(403)
    acct=current_account()
    # Class teachers may enter marks for every subject for their assigned class.
    # No per-subject teacher assignment is required.
    if val is not None and (val<0 or val>exam.max_marks): return jsonify({'error':f'Marks must be between 0 and {exam.max_marks}.'}),400
    m=Mark.query.filter_by(student_id=sid,subject_code=subj.code,exam_id=exam.id).first() or Mark(student_id=sid,subject_code=subj.code,exam_id=exam.id,max_marks=exam.max_marks)
    if m.locked: return jsonify({'error':'This mark is locked.'}),409
    m.marks=val; m.updated_by=current_account().username; m.max_marks=exam.max_marks; db.session.add(m); log_audit('mark_updated','Mark',m.id,{'student_id':sid,'exam':exam.name,'subject':subj.name,'marks':val}); db.session.commit(); return jsonify({'ok':True})

@app.post('/academics/import')
@staff_required
def academics_import():
    upload=request.files.get('file')
    if not upload or not upload.filename.lower().endswith(('.xlsx','.xlsm')):
        flash('Upload an Excel workbook (.xlsx).','error'); return redirect(url_for('academics'))
    try:
        from openpyxl import load_workbook
        wb=load_workbook(upload,read_only=True,data_only=True)
        ws=wb.active
        rows=list(ws.iter_rows(values_only=True))
        if len(rows)<2: raise ValueError('The workbook is empty.')
        headers=[str(x or '').strip().lower() for x in rows[0]]
        required={'admission','exam','subject','marks'}
        if not required.issubset(set(headers)): raise ValueError('Required columns: Admission, Exam, Subject, Marks')
        idx={h:i for i,h in enumerate(headers)}; changed=0
        for row in rows[1:]:
            admission=str(row[idx['admission']] or '').strip(); exam_name=str(row[idx['exam']] or '').strip(); subject_name=str(row[idx['subject']] or '').strip(); raw=row[idx['marks']]
            if not admission or not exam_name or not subject_name or raw in (None,''): continue
            student=Student.query.filter_by(admission_number=admission).first(); exam=Exam.query.filter(func.lower(Exam.name)==exam_name.lower()).first(); subject=Subject.query.filter(func.lower(Subject.name)==subject_name.lower()).first()
            if not student or not exam or not subject: continue
            value=float(raw)
            if value<0 or value>exam.max_marks: continue
            mark=Mark.query.filter_by(student_id=student.id,subject_code=subject.code,exam_id=exam.id).first() or Mark(student_id=student.id,subject_code=subject.code,exam_id=exam.id,max_marks=exam.max_marks)
            if mark.locked: continue
            mark.marks=value; mark.max_marks=exam.max_marks; mark.updated_by=current_account().username; db.session.add(mark); changed += 1
        log_audit('marks_excel_import','Academics',extra={'updated':changed,'filename':upload.filename}); db.session.commit(); flash(f'Imported {changed} mark(s).','success')
    except Exception as exc:
        db.session.rollback(); flash(f'Could not import workbook: {exc}','error')
    return redirect(url_for('academics'))

@app.context_processor
def template_helpers():
    acct = current_account()
    return {
        'subjects_for_class': subjects_for_class,
        'school_date': school_date,
        'school_time': school_time,
        'school_now': school_now,
        'clock_override_time': (db.session.get(SchoolClock,1).override_time if db.session.get(SchoolClock,1) else None),
        'grade_for_percent': grade_for_percent,
        'current_account': current_account,
        'profile_picture_url': profile_picture_url,
        'me': acct,
        'csrf_token': session.get('csrf',''),
    }

@app.get('/results')
@login_required()
def results():
    acct=current_account(); students=allowed_students_for_account(acct).order_by(Student.class_name,Student.section,Student.name).limit(1000).all(); exams=Exam.query.order_by(Exam.order_index).all(); ids=[s.id for s in students]
    marks=Mark.query.filter(Mark.student_id.in_(ids)).all() if ids else []
    by_student={}
    for m in marks: by_student.setdefault(m.student_id,[]).append(m)
    rows=[]
    for s in students:
        total=0; max_total=0; by_exam=[]; smarks=by_student.get(s.id,[])
        for e in exams:
            ms=[m for m in smarks if m.exam_id==e.id]; got=sum(m.marks or 0 for m in ms); mx=sum(m.max_marks for m in ms); pct=round(got/mx*100,2) if mx else 0
            total += got; max_total += mx; by_exam.append((e.name,got,mx,pct,grade_for_percent(pct)))
        pct=round(total/max_total*100,2) if max_total else 0
        rows.append({'student':s,'by_exam':by_exam,'percentage':pct,'grade':grade_for_percent(pct),'result':pass_fail_for_percent(pct)})
    setting=SchoolSetting.query.filter_by(key='academic_session').first(); school_session=setting.value if setting else os.getenv('ACADEMIC_SESSION','2026-27')
    return render_template('results.html',rows=rows,exams=exams,school_session=school_session)

@app.get('/results/export.xlsx')
@login_required()
def results_xlsx():
    rows=[]
    for s in allowed_students_for_account(current_account()).all():
        for e in Exam.query.order_by(Exam.order_index).all():
            ms=Mark.query.filter_by(student_id=s.id,exam_id=e.id).all(); rows.append({'Admission':s.admission_number,'Name':s.name,'Class':s.class_name,'Section':s.section,'Exam':e.name,'Marks':sum(m.marks or 0 for m in ms),'Max':sum(m.max_marks for m in ms)})
    return send_file(build_xlsx(rows,'Results'),as_attachment=True,download_name='results.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/admin/attendance/export.xlsx')
@staff_required
def attendance_xlsx():
    try:
        day=datetime.strptime(request.args.get('date',school_date().isoformat()),'%Y-%m-%d').date()
    except ValueError:
        day=school_date()
    cls=normalize_class(request.args.get('class_name','')); sec=normalize_section(request.args.get('section',''))
    q=allowed_students_for_account(current_account())
    if cls: q=q.filter(Student.class_name==cls)
    if sec: q=q.filter(Student.section==sec)
    students=q.order_by(Student.class_name,Student.section,Student.roll_number,Student.name).all()
    ids=[s.id for s in students]
    ats=Attendance.query.filter(Attendance.date==day,Attendance.student_id.in_(ids)).all() if ids else []
    by={a.student_id:a for a in ats}
    rows=[]
    for s in students:
        a=by.get(s.id)
        rows.append({'Roll':s.roll_number or '', 'Admission':s.admission_number, 'Name':s.name, 'Class':s.class_name, 'Section':s.section or '', 'Status':(a.status if a else 'absent').title(), 'Time':(a.time_in.strftime('%I:%M %p') if a and a.time_in else '')})
    return send_file(build_xlsx(rows,'Attendance Register'),as_attachment=True,download_name=f'attendance_{day.isoformat()}.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/admin/attendance/export.pdf')
@staff_required
def attendance_pdf():
    try:
        day=datetime.strptime(request.args.get('date',school_date().isoformat()),'%Y-%m-%d').date()
    except ValueError:
        day=school_date()
    cls=normalize_class(request.args.get('class_name','')); sec=normalize_section(request.args.get('section',''))
    q=allowed_students_for_account(current_account())
    if cls: q=q.filter(Student.class_name==cls)
    if sec: q=q.filter(Student.section==sec)
    students=q.order_by(Student.class_name,Student.section,Student.roll_number,Student.name).all()
    ids=[s.id for s in students]
    ats=Attendance.query.filter(Attendance.date==day,Attendance.student_id.in_(ids)).all() if ids else []
    by={a.student_id:a for a in ats}
    rows=[]
    for s in students:
        a=by.get(s.id)
        rows.append({'Roll':s.roll_number or '', 'Admission':s.admission_number, 'Name':s.name, 'Class':s.class_name, 'Section':s.section or '', 'Status':(a.status if a else 'absent').title(), 'Time':(a.time_in.strftime('%I:%M %p') if a and a.time_in else '')})
    return send_file(build_pdf(rows,'DAV PS KKP • Attendance Register',subtitle=day.strftime('%d.%m.%Y')),as_attachment=True,download_name=f'attendance_{day.isoformat()}.pdf',mimetype='application/pdf')

@app.get('/fees/export.xlsx')
@staff_required
def fees_xlsx():
    rows=[]
    for inv in FeeInvoice.query.order_by(FeeInvoice.created_at.desc()).all():
        paid=sum(p.amount for p in inv.payments)
        rows.append({'Student':inv.student.name,'Admission':inv.student.admission_number,'Fee':inv.title,'Due':inv.amount_due,'Paid':paid,'Balance':max(0,inv.amount_due-paid),'Status':inv.status,'Due Date':inv.due_date})
    return send_file(build_xlsx(rows,'Fee Ledger'),as_attachment=True,download_name='fee_ledger.xlsx',mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.get('/fees/export.pdf')
@staff_required
def fees_pdf():
    rows=[]
    for inv in FeeInvoice.query.order_by(FeeInvoice.created_at.desc()).all():
        paid=sum(p.amount for p in inv.payments)
        rows.append({'Student':inv.student.name,'Admission':inv.student.admission_number,'Fee':inv.title,'Due':inv.amount_due,'Paid':paid,'Balance':max(0,inv.amount_due-paid),'Status':inv.status})
    return send_file(build_pdf(rows,'DAV PS KKP • Fee Ledger'),as_attachment=True,download_name='fee_ledger.pdf',mimetype='application/pdf')

@app.get('/results/export.pdf')
@login_required()
def results_pdf():
    rows=[]
    for s in allowed_students_for_account(current_account()).all():
        for e in Exam.query.order_by(Exam.order_index).all():
            ms=Mark.query.filter_by(student_id=s.id,exam_id=e.id).all(); rows.append({'Admission':s.admission_number,'Name':s.name,'Class':s.class_name,'Section':s.section,'Exam':e.name,'Marks':sum(m.marks or 0 for m in ms),'Max':sum(m.max_marks for m in ms)})
    return send_file(build_pdf(rows,'Results'),as_attachment=True,download_name='results.pdf',mimetype='application/pdf')

def student_can_see_results(student):
    final=Exam.query.filter_by(is_final=True).first()
    if not final: return False
    pub=ResultPublication.query.filter_by(exam_id=final.id,class_name=student.class_name,section=student.section,published=True).first()
    return bool(pub)

@app.get('/my/attendance')
@login_required()
def my_attendance():
    acct=current_account(); rows=[]
    for s in allowed_students_for_account(acct).all():
        start,end=school_year_bounds(); today=min(school_date(),end)
        y=attendance_summary(s.id,start,today)
        term1_end=date(start.year,9,30); term2_start=date(start.year,10,1)
        t1=attendance_summary(s.id,start,min(today,term1_end))
        t2=attendance_summary(s.id,term2_start,today) if today>=term2_start else {'working_days':0,'present':0,'absent':0,'percentage':0}
        rows.append({'student':s,'year':y,'term1':t1,'term2':t2})
    return render_template('my_attendance.html',rows=rows)

@app.get('/my/results')
@login_required()
def my_results():
    acct=current_account(); data=[]
    for s in allowed_students_for_account(acct).all():
        data.append((s,student_can_see_results(s)))
    return render_template('my_results.html',rows=data)

@app.get('/teacher/signature')
@staff_required
def teacher_signature():
    t=Teacher.query.filter_by(account_id=current_account().id).first()
    if not t: abort(404)
    return render_template('teacher_signature.html',teacher=t)

@app.post('/teacher/signature')
@staff_required
def teacher_signature_save():
    t=Teacher.query.filter_by(account_id=current_account().id).first() or abort(404)

    # Teachers can either draw a signature on the canvas or upload an image.
    uploaded = request.files.get('signature_file')
    uploaded_value = ''
    if uploaded and uploaded.filename:
        raw = uploaded.read()
        if len(raw) > 2 * 1024 * 1024:
            flash('SIGNATURE IMAGE MUST BE 2 MB OR SMALLER.','error')
            return redirect(url_for('teacher_signature'))
        ext = Path(uploaded.filename).suffix.lower()
        mimetype = (uploaded.mimetype or '').lower()
        allowed = {'.png':'image/png', '.jpg':'image/jpeg', '.jpeg':'image/jpeg', '.webp':'image/webp'}
        if ext not in allowed or mimetype not in set(allowed.values()):
            flash('UPLOAD A PNG, JPG, JPEG, OR WEBP SIGNATURE IMAGE.','error')
            return redirect(url_for('teacher_signature'))
        uploaded_value = f"data:{allowed[ext]};base64,{base64.b64encode(raw).decode('ascii')}"

    drawn_value = request.form.get('signature_data','').strip()
    if uploaded_value:
        t.signature_data = uploaded_value[:200000]
    elif drawn_value.startswith('data:image/'):
        t.signature_data = drawn_value[:200000]
    else:
        flash('DRAW OR UPLOAD A SIGNATURE FIRST.','error')
        return redirect(url_for('teacher_signature'))

    log_audit('teacher_signature_updated','Teacher',t.id)
    db.session.commit()
    flash('Signature saved for future report cards.','success')
    return redirect(url_for('teacher_signature'))

@app.get('/admin/assignments')
@admin_required
def assignments():
    assignments=TeacherAssignment.query.order_by(TeacherAssignment.class_name,TeacherAssignment.section).all()
    teachers=Teacher.query.filter_by(active=True).order_by(Teacher.name).all()
    teacher_names={t.id:t.name for t in teachers}
    return render_template('assignments.html',teachers=teachers,assignments=assignments,teacher_names=teacher_names,class_options=class_list())

@app.post('/admin/assignments')
@admin_required
def assignments_save():
    tid=int(request.form.get('teacher_id')); cls=normalize_class(request.form.get('class_name')); sec=normalize_section(request.form.get('section'))
    teacher=db.session.get(Teacher,tid)
    if not teacher or not cls:
        flash('SELECT A VALID TEACHER AND CLASS.','error'); return redirect(url_for('assignments'))
    if sec and class_number(cls) < 9 and len(sec)>2:
        flash('SECTION MUST BE A SHORT CLASS SECTION SUCH AS A OR B.','error'); return redirect(url_for('assignments'))
    existing=TeacherAssignment.query.filter_by(class_name=cls,section=sec).first()
    if existing:
        flash('THIS CLASS/SECTION ALREADY HAS A CLASS TEACHER.','error'); return redirect(url_for('assignments'))
    x=TeacherAssignment(teacher_id=tid,class_name=cls,section=sec)
    db.session.add(x)
    log_audit('teacher_assignment_created','TeacherAssignment',extra={'teacher_id':tid,'class_name':cls,'section':sec})
    db.session.commit()
    flash('CLASS TEACHER ASSIGNMENT ADDED.','success')
    return redirect(url_for('assignments'))

@app.get('/api/subjects-for-class')
@admin_required
def subjects_for_class_api():
    cls=normalize_class(request.args.get('class_name','')); second=request.args.get('second_language'); third=request.args.get('third_language')
    return jsonify([{'code':c,'name':n} for c,n in subject_options_for_class(cls,second,third) if not c.endswith('_placeholder')])


def _json_dict(value):
    try: return json.loads(value or '{}')
    except Exception: return {}


def _can_edit_report_card(acct, student):
    if acct.role == 'admin': return True
    if acct.role != 'teacher': return False
    teacher=Teacher.query.filter_by(account_id=acct.id,active=True).first()
    if not teacher: return False
    return TeacherAssignment.query.filter_by(teacher_id=teacher.id,class_name=student.class_name,section=student.section).first() is not None

@app.route('/report-card/config/<int:sid>', methods=['GET','POST'])
@staff_required
def report_card_config(sid):
    student=db.session.get(Student,sid) or abort(404); acct=current_account()
    if not _can_edit_report_card(acct, student): abort(403)
    setting=SchoolSetting.query.filter_by(key='academic_session').first(); session_name=setting.value if setting else os.getenv('ACADEMIC_SESSION','2026-27')
    cfg=ReportCardConfig.query.filter_by(student_id=sid).first()
    if not cfg:
        cfg=ReportCardConfig(student_id=sid,academic_session=session_name,next_academic_session='')
        ta=TeacherAssignment.query.filter_by(class_name=student.class_name,section=student.section).first()
        if ta:
            tt=Teacher.query.get(ta.teacher_id); cfg.class_teacher_name=tt.name if tt else ''
        db.session.add(cfg); db.session.flush()
    subjects=subjects_for_class(student.class_name,student.second_language,student.third_language)
    if request.method=='POST':
        def clean_text(name,limit=4000): return str(request.form.get(name,'')).strip()[:limit]
        cfg.academic_session=clean_text('academic_session',30) or session_name
        cfg.house=clean_text('house',80); cfg.class_teacher_name=clean_text('class_teacher_name',160)
        cfg.remarks=clean_text('remarks'); cfg.principal_remarks=clean_text('principal_remarks'); cfg.date_result=clean_text('date_result',80)
        cfg.next_academic_session=clean_text('next_academic_session',30); cfg.session_begins=clean_text('session_begins',40); cfg.summer_break_from=clean_text('summer_break_from',40); cfg.school_reopens=clean_text('school_reopens',40)
        co={k:request.form.get(k,'').strip() for k in ['work_education','art_education','health_physical']}
        dis={k:request.form.get(k,'').strip() for k in ['discipline','regularity','punctuality']}
        health={k:request.form.get(k,'').strip() for k in ['term1_height','term1_weight','term2_height','term2_weight']}
        layout={k:request.form.get(k,'').strip()[:200] for k in ['report_title','attendance_title','scholastic_title','development_title','guide_title','class_details_title','pass_rule','teacher_signature_label','principal_signature_label','parent_signature_label']}
        cfg.co_scholastic_json=json.dumps(co); cfg.discipline_json=json.dumps(dis); cfg.health_json=json.dumps(health); cfg.layout_json=json.dumps(layout); cfg.updated_by=acct.username
        for subj in subjects:
            component=AssessmentComponent.query.filter_by(student_id=sid,subject_code=subj.code).first() or AssessmentComponent(student_id=sid,subject_code=subj.code)
            for field in ['multiple_assessment','subject_enrichment','portfolio']:
                raw=request.form.get(f'{subj.code}_{field}','')
                try: setattr(component,field,max(0,min(5,float(raw))) if raw.strip() else 0)
                except ValueError: setattr(component,field,0)
            # internal is computed from best PT + the three 5-point components
            db.session.add(component)
        log_audit('report_card_config_updated','ReportCardConfig',cfg.id,{'student_id':sid}); db.session.commit(); flash('REPORT CARD DETAILS SAVED.','success'); return redirect(url_for('results'))
    components={x.subject_code:x for x in AssessmentComponent.query.filter_by(student_id=sid).all()}
    layout=_json_dict(getattr(cfg,'layout_json','{}'))
    return render_template('report_card_config.html',student=student,cfg=cfg,co=_json_dict(cfg.co_scholastic_json),dis=_json_dict(cfg.discipline_json),health=_json_dict(cfg.health_json),layout=layout,subjects=subjects,components=components,session_name=session_name)

@app.get('/report-card/<int:sid>.pdf')
@login_required()
def report_card(sid):
    s=db.session.get(Student,sid); acct=current_account()
    if not s or allowed_students_for_account(acct).filter_by(id=sid).first() is None: abort(403)
    if acct.role in {'student','parent'} and not student_can_see_results(s):
        return render_template('locked_result.html',student=s), 403
    setting=SchoolSetting.query.filter_by(key='academic_session').first(); session_name=setting.value if setting else os.getenv('ACADEMIC_SESSION','2026-27')
    cfg=ReportCardConfig.query.filter_by(student_id=sid).first()
    config={}
    if cfg:
        config={'house':cfg.house,'class_teacher_name':cfg.class_teacher_name,'co_scholastic':_json_dict(cfg.co_scholastic_json),'discipline':_json_dict(cfg.discipline_json),'health':_json_dict(cfg.health_json),'remarks':cfg.remarks or '','principal_remarks':cfg.principal_remarks or '','date_result':cfg.date_result or '','next_academic_session':cfg.next_academic_session or '','session_begins':cfg.session_begins or '','summer_break_from':cfg.summer_break_from or '','school_reopens':cfg.school_reopens or '','layout':_json_dict(getattr(cfg,'layout_json','{}'))}
    start,end=school_year_bounds(); today=min(school_date(),end); t1_end=date(start.year,9,30); t2_start=date(start.year,10,1)
    att=attendance_summary(s.id,start,today); t1=attendance_summary(s.id,start,min(today,t1_end)); t2=attendance_summary(s.id,t2_start,today) if today>=t2_start else {'working_days':0,'present':0,'absent':0,'percentage':0}
    exams=Exam.query.order_by(Exam.order_index).all(); subjects=subjects_for_class(s.class_name,s.second_language,s.third_language)
    smarks=Mark.query.filter_by(student_id=s.id).all(); marks_by_subject={(m.subject_code,m.exam_id):m for m in smarks}
    comps={x.subject_code:{'multiple_assessment':x.multiple_assessment or 0,'subject_enrichment':x.subject_enrichment or 0,'portfolio':x.portfolio or 0} for x in AssessmentComponent.query.filter_by(student_id=s.id).all()}
    teacher_signature=None
    ta=TeacherAssignment.query.filter_by(class_name=s.class_name,section=s.section).first()
    if ta:
        teacher=Teacher.query.get(ta.teacher_id); teacher_signature=teacher.signature_data if teacher else None
    logo_path=BASE_DIR/'static'/'branding'/'dav-ps-kkp-logo.png'
    buf=build_report_card(s,session_name,att,t1,t2,subjects,exams,marks_by_subject,teacher_signature,config,logo_path,comps)
    return send_file(buf,as_attachment=True,download_name=f'report_card_{s.admission_number}.pdf',mimetype='application/pdf')

@app.route('/admin/accounts', methods=['GET','POST'])
@admin_required
def admin_accounts():
    if request.method=='POST':
        role=request.form.get('role'); username=request.form.get('username','').strip(); password=request.form.get('password',''); name=normalize_school_name(request.form.get('name'))
        child_admissions=[x.strip().upper() for x in request.form.get('child_admissions','').replace('\n',',').split(',') if x.strip()]
        child_ids=[s.id for s in Student.query.filter(Student.admission_number.in_(child_admissions)).all()]
        if role not in {'admin','teacher','parent'} or not username or len(password)<8 or not name:
            flash('Name, username, role and 8+ character temporary password are required. Student accounts are created only through registration.','error')
        elif role=='parent' and not child_ids:
            flash('A parent account must be linked to at least one student.','error')
        elif Account.query.filter(func.lower(Account.username)==username.lower()).first():
            flash('Username already exists.','error')
        else:
            acct=Account(username=username,password_hash=generate_password_hash(password),role=role,display_name=name,must_change_password=True)
            db.session.add(acct); db.session.flush()
            if role=='parent':
                parent=Parent(name=name,account_id=acct.id,phone=request.form.get('phone','').strip(),email=request.form.get('email','').strip())
                db.session.add(parent); db.session.flush()
                for sid in child_ids: db.session.add(ParentStudent(parent_id=parent.id,student_id=sid))
            elif role=='teacher':
                db.session.add(Teacher(name=name,phone=request.form.get('phone','').strip(),email=request.form.get('email','').strip(),account_id=acct.id))
            log_audit('account_created','Account',acct.id,{'role':role,'children':child_ids}); db.session.commit(); flash(f'{role.title()} account created.','success')

    role=request.args.get('role','').lower(); search=request.args.get('q','').strip(); page=max(1,request.args.get('page',1,type=int)); per_page=25
    # Admins
    admin_items=[]; admin_count=0
    if not role or role=='admin':
        aq=Account.query.filter_by(role='admin',active=True)
        if search: aq=aq.filter(or_(Account.username.ilike(f'%{search}%'),Account.display_name.ilike(f'%{search}%')))
        admin_count=aq.count(); admin_items=aq.order_by(Account.display_name).offset((page-1)*per_page).limit(per_page).all()

    # Teachers, with assignment strings loaded for only the current page.
    teacher_items=[]; teacher_count=0; assignment_map={}; teacher_accounts={}
    if not role or role=='teacher':
        tq=Teacher.query.filter_by(active=True)
        if search:
            tq=tq.outerjoin(Account,Teacher.account_id==Account.id).filter(or_(Teacher.name.ilike(f'%{search}%'),Account.username.ilike(f'%{search}%')))
        teacher_count=tq.count(); teacher_items=tq.order_by(Teacher.name).offset((page-1)*per_page).limit(per_page).all()
        tids=[t.id for t in teacher_items]
        for t in teacher_items: teacher_accounts[t.id]=t.account
        if tids:
            assigns=TeacherAssignment.query.filter(TeacherAssignment.teacher_id.in_(tids)).order_by(TeacherAssignment.class_name,TeacherAssignment.section).all()
            for a in assigns: assignment_map.setdefault(a.teacher_id,[]).append(f'{a.class_name}-{a.section}' if a.section else a.class_name)

    # Students are sourced from Student, not Account, so every registered student is visible even if an old record is missing account_id.
    student_items=[]; student_count=0; parent_count_by_student={}
    if not role or role=='student':
        sq=Student.query.filter_by(active=True)
        if search:
            sq=sq.outerjoin(Account,Student.account_id==Account.id).filter(or_(Student.name.ilike(f'%{search}%'),Student.admission_number.ilike(f'%{search}%'),Student.roll_number.ilike(f'%{search}%'),Account.username.ilike(f'%{search}%')))
        student_count=sq.count(); student_items=sq.order_by(Student.class_name,Student.section,Student.roll_number,Student.name).offset((page-1)*per_page).limit(per_page).all()
        sids=[st.id for st in student_items]
        if sids:
            links=ParentStudent.query.filter(ParentStudent.student_id.in_(sids)).all()
            for link in links:
                parent_count_by_student[link.student_id]=parent_count_by_student.get(link.student_id,0)+1

    # Parents are sourced from Parent so they are always visible and their linked children remain explicit.
    parent_items=[]; parent_count=0; child_map={}
    if not role or role=='parent':
        pq=Parent.query.filter(Parent.account_id.isnot(None))
        if search:
            pq=pq.outerjoin(Account,Parent.account_id==Account.id).filter(or_(Parent.name.ilike(f'%{search}%'),Account.username.ilike(f'%{search}%')))
        parent_count=pq.count(); parent_items=pq.order_by(Parent.name).offset((page-1)*per_page).limit(per_page).all()
        pids=[p.id for p in parent_items]
        if pids:
            links=ParentStudent.query.filter(ParentStudent.parent_id.in_(pids)).all()
            sids=[x.student_id for x in links]
            smap={st.id:st for st in Student.query.filter(Student.id.in_(sids)).all()} if sids else {}
            for link in links:
                st=smap.get(link.student_id)
                if st: child_map.setdefault(link.parent_id,[]).append(st)

    totals={'admin':admin_count,'teacher':teacher_count,'student':student_count,'parent':parent_count}
    return render_template('accounts.html',role=role,search=search,page=page,per_page=per_page,
        admin_items=admin_items,teacher_items=teacher_items,student_items=student_items,parent_items=parent_items,
        totals=totals,assignment_map=assignment_map,child_map=child_map,teacher_accounts=teacher_accounts,parent_count_by_student=parent_count_by_student)

FEE_CLASS_GROUPS = ['Nursery & LKG','UKG to IV','V to IX','X','XI & XII']
FEE_TERM_LABELS = [('term_i','TERM-I'),('term_ii','TERM-II'),('term_iii','TERM-III'),('term_iv','TERM-IV')]
DEFAULT_FEE_PAYMENT_WINDOWS_2026_27 = {
    'term_i': ('1st Term', date(2026,4,16), date(2026,6,25), date(2026,6,26)),
    'term_ii': ('2nd Term', date(2026,7,1), date(2026,9,25), date(2026,9,26)),
    'term_iii': ('3rd Term', date(2026,10,1), date(2026,12,25), date(2026,12,26)),
    'term_iv': ('4th Term', date(2027,1,1), date(2027,2,20), date(2027,2,21)),
}
DEFAULT_FEE_STRUCTURE_2026_27 = {
    'Nursery & LKG': (19700,19700,19700,0),
    'UKG to IV': (14775,14775,14775,14775),
    'V to IX': (15480,15480,15480,15480),
    'X': (16395,16395,16395,16395),
    'XI & XII': (17700,17700,17700,17700),
}

def fee_group_for_student(student):
    raw=str(student.class_name or '').strip().upper()
    if raw in {'NURSERY','LKG'}: return 'Nursery & LKG'
    if raw=='UKG' or raw in {'1','2','3','4','I','II','III','IV'}: return 'UKG to IV'
    try:
        n=int(raw)
        if 1 <= n <= 4: return 'UKG to IV'
        if 5 <= n <= 9: return 'V to IX'
        if n==10: return 'X'
        if n in (11,12): return 'XI & XII'
    except Exception: pass
    if raw in {'I','II','III','IV'}: return 'UKG to IV'
    if raw in {'V','VI','VII','VIII','IX'}: return 'V to IX'
    if raw=='X': return 'X'
    if raw in {'XI','XII'}: return 'XI & XII'
    return None

def ensure_fee_structure_seed(session_name='2026-27'):
    changed=False
    for group, terms in DEFAULT_FEE_STRUCTURE_2026_27.items():
        row=FeeStructure.query.filter_by(academic_session=session_name,class_group=group).first()
        if not row:
            row=FeeStructure(academic_session=session_name,class_group=group,term_i=terms[0],term_ii=terms[1],term_iii=terms[2],term_iv=terms[3],total=sum(terms),updated_by='system')
            db.session.add(row); changed=True
    if changed: db.session.commit()

def ensure_fee_payment_windows(session_name='2026-27'):
    defaults = DEFAULT_FEE_PAYMENT_WINDOWS_2026_27 if session_name == '2026-27' else {}
    changed=False
    for key, (label, start, end, fine_from) in defaults.items():
        row=FeePaymentWindow.query.filter_by(academic_session=session_name, term_key=key).first()
        if not row:
            row=FeePaymentWindow(academic_session=session_name, term_key=key, term_label=label, payment_start=start, payment_end=end, fine_from=fine_from, updated_by='system')
            db.session.add(row); changed=True
    if changed:
        db.session.commit()


def fee_structure_for_student(student, session_name='2026-27'):
    group=fee_group_for_student(student)
    return FeeStructure.query.filter_by(academic_session=session_name,class_group=group).first() if group else None

def fee_term_amount(row, term):
    return float(getattr(row, term, 0) or 0) if row else 0.0

@app.get('/api/fees/students/search')
@admin_required
def fee_student_search():
    q=' '.join((request.args.get('q') or '').split())
    if len(q)<1: return jsonify([])
    like=f'%{q}%'
    query=Student.query.filter(Student.active==True)
    query=query.filter(or_(Student.admission_number.ilike(like),Student.name.ilike(like),Student.roll_number.ilike(like),Student.class_name.ilike(like),Student.section.ilike(like)))
    rows=query.order_by(Student.name.asc()).limit(30).all()
    out=[]
    for st in rows:
        fs=fee_structure_for_student(st, request.args.get('session','2026-27'))
        out.append({'id':st.id,'name':st.name,'admission_number':st.admission_number,'roll_number':st.roll_number or '', 'class_name':st.class_name,'section':st.section or '', 'fee_group':fee_group_for_student(st) or '', 'fee_structure':{'term_i':fee_term_amount(fs,'term_i'),'term_ii':fee_term_amount(fs,'term_ii'),'term_iii':fee_term_amount(fs,'term_iii'),'term_iv':fee_term_amount(fs,'term_iv'),'total':fee_term_amount(fs,'total')}})
    return jsonify(out)

@app.route('/admin/fees/structure', methods=['GET','POST'])
@admin_required
def fee_structure_admin():
    ensure_fee_schema()
    if request.method=='POST':
        session_name=(request.form.get('academic_session') or '2026-27').strip()[:30]
        for group in FEE_CLASS_GROUPS:
            vals=[]
            for key in ['term_i','term_ii','term_iii','term_iv']:
                raw=(request.form.get(f'{group}__{key}') or '0').replace(',','').strip()
                try: vals.append(max(0,float(raw or 0)))
                except ValueError: vals.append(0.0)
            row=FeeStructure.query.filter_by(academic_session=session_name,class_group=group).first()
            if not row: row=FeeStructure(academic_session=session_name,class_group=group)
            row.academic_session=session_name; row.class_group=group; row.term_i,row.term_ii,row.term_iii,row.term_iv=vals; row.total=sum(vals); row.updated_by=current_account().username; db.session.add(row)
        term_window_labels={'term_i':'1st Term','term_ii':'2nd Term','term_iii':'3rd Term','term_iv':'4th Term'}
        for key, label in term_window_labels.items():
            start_raw=(request.form.get(f'{key}__start') or '').strip()
            end_raw=(request.form.get(f'{key}__end') or '').strip()
            fine_raw=(request.form.get(f'{key}__fine') or '').strip()
            if not (start_raw and end_raw and fine_raw):
                flash(f'Enter all payment dates for {label}.','error')
                return redirect(url_for('fee_structure_admin',session=session_name))
            try:
                start_dt=datetime.strptime(start_raw,'%Y-%m-%d').date()
                end_dt=datetime.strptime(end_raw,'%Y-%m-%d').date()
                fine_dt=datetime.strptime(fine_raw,'%Y-%m-%d').date()
                if end_dt < start_dt or fine_dt <= end_dt:
                    raise ValueError
            except ValueError:
                flash(f'Invalid payment dates for {label}. Fine date must be after the payment window.','error')
                return redirect(url_for('fee_structure_admin',session=session_name))
            row=FeePaymentWindow.query.filter_by(academic_session=session_name,term_key=key).first()
            if not row: row=FeePaymentWindow(academic_session=session_name,term_key=key,term_label=label)
            row.payment_start=start_dt; row.payment_end=end_dt; row.fine_from=fine_dt; row.updated_by=current_account().username
            db.session.add(row)

        upload=request.files.get('fee_document')
        if upload and upload.filename:
            raw=upload.read()
            if len(raw)>12*1024*1024:
                flash('Fee notice must be 12 MB or smaller.','error'); return redirect(url_for('fee_structure_admin',session=session_name))
            ext=Path(upload.filename).suffix.lower(); allowed={'.pdf':'application/pdf','.png':'image/png','.jpg':'image/jpeg','.jpeg':'image/jpeg','.webp':'image/webp'}
            mime=allowed.get(ext,(upload.mimetype or 'application/octet-stream'))
            if ext not in allowed and not (upload.mimetype or '').startswith('image/') and upload.mimetype!='application/pdf':
                flash('Upload a PDF, PNG, JPG, JPEG or WEBP fee notice.','error'); return redirect(url_for('fee_structure_admin',session=session_name))
            db.session.add(FeeStructureDocument(academic_session=session_name,filename=upload.filename[:255],mimetype=mime,file_data=raw,uploaded_by=current_account().username))
        log_audit('fee_structure_update','FeeStructure',extra={'academic_session':session_name}); db.session.commit(); flash(f'Fee structure saved for {session_name}.','success'); return redirect(url_for('fee_structure_admin',session=session_name))
    session_name=(request.args.get('session') or '2026-27').strip()[:30]
    ensure_fee_structure_seed(session_name)
    ensure_fee_payment_windows(session_name)
    rows=FeeStructure.query.filter_by(academic_session=session_name).order_by(FeeStructure.id.asc()).all()
    windows=FeePaymentWindow.query.filter_by(academic_session=session_name).order_by(FeePaymentWindow.id.asc()).all()
    docs=FeeStructureDocument.query.filter_by(academic_session=session_name).order_by(FeeStructureDocument.uploaded_at.desc()).limit(10).all()
    return render_template('fee_structure.html',session_name=session_name,rows={r.class_group:r for r in rows},groups=FEE_CLASS_GROUPS,terms=FEE_TERM_LABELS,windows=windows,docs=docs)

@app.get('/admin/fees/document/<int:doc_id>')
@admin_required
def fee_structure_document(doc_id):
    doc=db.session.get(FeeStructureDocument,doc_id) or abort(404)
    return send_file(BytesIO(doc.file_data),download_name=doc.filename,mimetype=doc.mimetype,as_attachment=False)

@app.route('/fees', methods=['GET','POST'])
@admin_required
def fees():
    ensure_fee_schema()
    ensure_fee_structure_seed('2026-27')
    ensure_fee_payment_windows('2026-27')
    if request.method=='POST':
        try:
            sid=int(request.form.get('student_id') or 0)
            session_name=(request.form.get('academic_session') or '2026-27').strip()[:30]
            term=request.form.get('term') or 'annual'
            due=request.form.get('due_date') or None
            student=db.session.get(Student,sid) or abort(404)
            structure=fee_structure_for_student(student,session_name)
            if not structure:
                flash(f'No fee structure is configured for {student.class_name} for {session_name}.','error'); return redirect(url_for('fees'))
            if term=='annual': amount=float(structure.total); title=f'School Fee {session_name} • Annual'
            elif term in {'term_i','term_ii','term_iii','term_iv'}:
                amount=fee_term_amount(structure,term); label=dict(FEE_TERM_LABELS)[term]; title=f'School Fee {session_name} • {label}'
            else:
                raise ValueError('Invalid fee term')
            if amount<=0:
                flash('That fee term has no amount configured.','error'); return redirect(url_for('fees'))
            existing=FeeInvoice.query.filter_by(student_id=sid,title=title).first()
            if existing:
                flash('That fee invoice already exists for this student.','error'); return redirect(url_for('fees'))
            inv=FeeInvoice(student_id=sid,title=title,amount_due=amount,due_date=datetime.strptime(due,'%Y-%m-%d').date() if due else None,status='pending'); db.session.add(inv); log_audit('fee_invoice_created','FeeInvoice',extra={'student_id':sid,'amount':amount,'session':session_name,'term':term}); db.session.commit(); flash('Fee invoice created from the configured class fee structure.','success'); return redirect(url_for('fees'))
        except (TypeError,ValueError):
            db.session.rollback(); flash('Invalid fee request. Select a valid student and term.','error'); return redirect(url_for('fees'))
    rows=FeeInvoice.query.order_by(FeeInvoice.created_at.desc()).limit(200).all()
    return render_template('fees.html',rows=rows,session_name='2026-27')

@app.post('/fees/<int:invoice_id>/pay')
@admin_required
def fee_pay(invoice_id):
    inv=db.session.get(FeeInvoice,invoice_id) or abort(404); paid=sum(x.amount for x in inv.payments); remaining=max(0,inv.amount_due-paid); amount=float(request.form.get('amount') or remaining)
    if amount<=0 or amount>remaining: flash('Payment amount is invalid.','error'); return redirect(url_for('fees'))
    rec=f'RCPT-{datetime.utcnow().strftime("%Y%m%d%H%M%S")}-{inv.id}'
    while FeePayment.query.filter_by(receipt_no=rec).first(): rec += 'X'
    db.session.add(FeePayment(invoice_id=inv.id,amount=amount,method=request.form.get('method','offline'),receipt_no=rec,received_by=current_account().username))
    inv.status='paid' if amount>=remaining else 'partial'; log_audit('fee_payment_recorded','FeeInvoice',inv.id,{'amount':amount,'receipt':rec}); db.session.commit(); flash('Payment recorded.','success'); return redirect(url_for('fees'))

@app.get('/my/fees')
@login_required()
def my_fees():
    if current_account().role not in {'parent','admin'}: abort(403)
    students=allowed_students_for_account(current_account()).all(); ids=[s.id for s in students]; rows=FeeInvoice.query.filter(FeeInvoice.student_id.in_(ids)).order_by(FeeInvoice.created_at.desc()).all() if ids else []
    return render_template('my_fees.html',rows=rows)


# ---------------------------------------------------------------------------
# School calendar (restored from the original attendance system, adapted to
# the current account/role model)
# ---------------------------------------------------------------------------
_MONTH_ALIASES = {
    'jan':1,'january':1,'feb':2,'february':2,'mar':3,'march':3,'apr':4,'april':4,
    'may':5,'jun':6,'june':6,'jul':7,'july':7,'aug':8,'august':8,'sep':9,
    'sept':9,'september':9,'oct':10,'october':10,'nov':11,'november':11,
    'dec':12,'december':12,
}
_DATE_PATTERNS = [
    re.compile(r"\\b(?P<y>20\\d{2})[-/.](?P<m>\\d{1,2})[-/.](?P<d>\\d{1,2})\\b"),
    re.compile(r"\\b(?P<d>\\d{1,2})[-/.](?P<m>\\d{1,2})[-/.](?P<y>20\\d{2})\\b"),
    re.compile(r"\\b(?P<d>\\d{1,2})\\s+(?P<m>[A-Za-z]{3,9}),?\\s+(?P<y>20\\d{2})\\b", re.I),
    re.compile(r"\\b(?P<m>[A-Za-z]{3,9})\\s+(?P<d>\\d{1,2})(?:st|nd|rd|th)?,?\\s+(?P<y>20\\d{2})\\b", re.I),
]
_NONWORKING = ('holiday','holidays','vacation','vacations','closed','non-working','non working','no school','school holiday','festival','break','leave')
_WORKING = ('working day','working days','school day','school days','instruction day','instruction days','open','working')

def get_working_days_setting():
    try:
        return 5 if int(os.getenv('WORKING_DAYS','6')) == 5 else 6
    except (TypeError,ValueError):
        return 6

def weekly_default_is_working(day):
    return day.weekday() < get_working_days_setting()

def is_working_day(day):
    override=SchoolCalendar.query.filter_by(date=day).first()
    return bool(override.is_working) if override else weekly_default_is_working(day)

def _calendar_date_from_match(match, fallback_year=None):
    v=match.groupdict(); raw_year=v.get('y') or fallback_year
    try:
        year=int(raw_year); mraw=v.get('m'); month=int(mraw) if str(mraw).isdigit() else _MONTH_ALIASES.get(str(mraw).lower()); day=int(v.get('d'))
        if month is None or not 1<=month<=12 or not 1<=day<=31: return None
        return datetime(year,month,day).date()
    except (TypeError,ValueError):
        return None

def parse_calendar_text(text_value):
    text_value=re.sub(r"[\u00a0\r]+"," ",text_value or '')
    text_value=re.sub(r"[ \t]+"," ",text_value)
    lines=[line.strip() for line in text_value.split('\n') if line.strip()]
    years=[int(x) for x in re.findall(r"\b(20\d{2})\b",text_value)]
    fallback_year=years[0] if years else school_date().year
    detected={}
    for i,line in enumerate(lines):
        matches=[]
        for pattern in _DATE_PATTERNS: matches.extend(pattern.finditer(line))
        if not matches: continue
        context=' '.join(lines[max(0,i-1):i+1]); low=context.lower()
        non_working=any(w in low for w in _NONWORKING); working=any(w in low for w in _WORKING)
        proposed=False if non_working else True if working else None
        reason='Holiday / non-working' if non_working else 'Working day' if working else 'Detected date'
        for match in matches:
            day=_calendar_date_from_match(match,fallback_year)
            if not day: continue
            if day in detected and detected[day]['is_working'] is not None and proposed is None: continue
            detected[day]={'date':day.isoformat(),'is_working':proposed,'reason':reason}
    for day,item in detected.items():
        if item['is_working'] is None:
            d=datetime.strptime(day,'%Y-%m-%d').date(); item['is_working']=weekly_default_is_working(d); item['reason']='Imported; weekly default used'
    return sorted(detected.values(),key=lambda x:x['date'])

def extract_calendar_upload_text(storage):
    filename=(storage.filename or '').lower(); raw=storage.read()
    if not raw: raise ValueError('The calendar file is empty.')
    if filename.endswith('.pdf') or storage.mimetype=='application/pdf':
        if PdfReader is None: raise ValueError('PDF import is unavailable because pypdf is not installed.')
        reader=PdfReader(io.BytesIO(raw)); text_value='\n'.join((page.extract_text() or '') for page in reader.pages).strip()
        if not text_value: raise ValueError('This PDF contains no selectable text. Try an image or text-based PDF.')
        return text_value
    if filename.endswith(('.png','.jpg','.jpeg','.webp')) or storage.mimetype.startswith('image/'):
        if pytesseract is None: raise ValueError('Image OCR is unavailable. Install pytesseract/Tesseract or upload a PDF.')
        import numpy as np, cv2
        arr=np.frombuffer(raw,dtype=np.uint8); frame=cv2.imdecode(arr,cv2.IMREAD_COLOR)
        if frame is None: raise ValueError('The calendar image could not be decoded.')
        gray=cv2.cvtColor(frame,cv2.COLOR_BGR2GRAY)
        if max(gray.shape[:2])<1600:
            scale=1600/max(gray.shape[:2]); gray=cv2.resize(gray,None,fx=scale,fy=scale,interpolation=cv2.INTER_CUBIC)
        return pytesseract.image_to_string(gray,config='--psm 6')
    raise ValueError('Upload a PDF, PNG, JPG, JPEG or WEBP school calendar.')

@app.get('/admin/calendar')
@admin_required
def admin_calendar_page():
    return render_template('calendar.html',calendar_readonly=False)

@app.get('/teacher/calendar')
@staff_required
def teacher_calendar_page():
    if current_account().role != 'teacher': abort(403)
    return render_template('calendar.html',calendar_readonly=True)

@app.get('/calendar')
@login_required()
def calendar_view_redirect():
    if current_account().role=='admin': return redirect(url_for('admin_calendar_page'))
    if current_account().role=='teacher': return redirect(url_for('teacher_calendar_page'))
    return render_template('calendar.html',calendar_readonly=True)

@app.get('/api/calendar')
@login_required()
def calendar_api():
    year=request.args.get('year',school_date().year,type=int); month=request.args.get('month',school_date().month,type=int)
    if not 1<=month<=12 or not 1900<=year<=2200: return jsonify({'error':'Invalid calendar month'}),400
    start=datetime(year,month,1).date(); nxt=datetime(year+1,1,1).date() if month==12 else datetime(year,month+1,1).date(); end=nxt-timedelta(days=1)
    overrides={r.date:r for r in SchoolCalendar.query.filter(SchoolCalendar.date>=start,SchoolCalendar.date<=end).all()}
    rows=[]; day=start
    while day<=end:
        ov=overrides.get(day); rows.append({'date':day.isoformat(),'is_working':bool(ov.is_working) if ov else weekly_default_is_working(day),'reason':ov.reason if ov else '','override':bool(ov)}); day+=timedelta(days=1)
    return jsonify(rows)

@app.post('/api/calendar')
@admin_required
def calendar_set():
    data=request.get_json(silent=True) or {}; raw=str(data.get('date') or '')
    try: day=datetime.strptime(raw,'%Y-%m-%d').date()
    except ValueError: return jsonify({'error':'Valid date is required'}),400
    row=SchoolCalendar.query.filter_by(date=day).first() or SchoolCalendar(date=day)
    row.is_working=bool(data.get('is_working')); row.reason=(data.get('reason') or '').strip()[:255] or None; db.session.add(row); db.session.flush(); log_audit('calendar_update','SchoolCalendar',row.id,{'date':day.isoformat(),'is_working':row.is_working,'reason':row.reason}); db.session.commit(); return jsonify({'message':'Calendar updated'})

@app.post('/api/calendar/reset')
@admin_required
def calendar_reset():
    data=request.get_json(silent=True) or {}; raw=str(data.get('date') or '')
    try: day=datetime.strptime(raw,'%Y-%m-%d').date()
    except ValueError: return jsonify({'error':'Valid date is required'}),400
    row=SchoolCalendar.query.filter_by(date=day).first()
    if row:
        rid=row.id; db.session.delete(row); log_audit('calendar_reset','SchoolCalendar',rid,{'date':day.isoformat()}); db.session.commit()
    return jsonify({'date':day.isoformat(),'is_working':weekly_default_is_working(day)})

@app.post('/api/calendar/bulk')
@admin_required
def calendar_bulk():
    data=request.get_json(silent=True) or {}; dates=data.get('dates') or []; is_working=bool(data.get('is_working')); reason=(data.get('reason') or '').strip()[:255] or None; changed=0
    try:
        for raw in dates:
            day=datetime.strptime(str(raw),'%Y-%m-%d').date(); row=SchoolCalendar.query.filter_by(date=day).first() or SchoolCalendar(date=day); row.is_working=is_working; row.reason=reason; db.session.add(row); changed+=1
        log_audit('calendar_bulk_update','SchoolCalendar',extra={'changed':changed,'is_working':is_working,'reason':reason}); db.session.commit()
    except (TypeError,ValueError): db.session.rollback(); return jsonify({'error':'Every date must use YYYY-MM-DD'}),400
    return jsonify({'message':f'Updated {changed} calendar date(s)'})

@app.post('/api/calendar/import')
@admin_required
def calendar_import():
    upload=request.files.get('calendar_file'); uploaded_text=(request.form.get('calendar_text') or '').strip()
    if not upload and not uploaded_text: return jsonify({'error':'Choose a school calendar PDF or image first.'}),400
    try:
        text_value=uploaded_text or extract_calendar_upload_text(upload); rows=parse_calendar_text(text_value)
        if not rows: return jsonify({'error':'No dates could be detected. You can also edit days manually.'}),400
        return jsonify({'dates':rows,'working_count':sum(1 for r in rows if r['is_working']),'non_working_count':sum(1 for r in rows if not r['is_working']),'text_preview':text_value[:2000]})
    except ValueError as exc: return jsonify({'error':str(exc)}),400
    except Exception as exc:
        app.logger.exception('Calendar import failed'); return jsonify({'error':f'Calendar import failed: {exc}'}),500

@app.post('/api/calendar/import/apply')
@admin_required
def calendar_import_apply():
    rows=(request.get_json(silent=True) or {}).get('dates') or []; changed=0
    try:
        for item in rows:
            day=datetime.strptime(str(item.get('date') or ''),'%Y-%m-%d').date(); row=SchoolCalendar.query.filter_by(date=day).first() or SchoolCalendar(date=day); row.is_working=bool(item.get('is_working')); row.reason=(item.get('reason') or '').strip()[:255] or None; db.session.add(row); changed+=1
        log_audit('calendar_import','SchoolCalendar',extra={'changed':changed}); db.session.commit()
    except (TypeError,ValueError): db.session.rollback(); return jsonify({'error':'One or more detected calendar dates are invalid.'}),400
    return jsonify({'message':f'Applied {changed} calendar date(s)'})

@app.get('/audit')
@audit_network_required()
def audit_login():
    if session.get('audit_auth'): return redirect(url_for('audit_view'))
    return render_template('audit_login.html')

@app.post('/audit')
@audit_network_required()
def audit_auth():
    if secrets.compare_digest(request.form.get('username',''),os.getenv('AUDIT_USERNAME','auditor')) and secrets.compare_digest(request.form.get('password',''),os.getenv('AUDIT_PASSWORD','ChangeAuditPasswordImmediately!')):
        session['audit_auth']=True; session['audit_username']=request.form.get('username'); return redirect(url_for('audit_view'))
    return render_template('audit_login.html',error='Invalid audit credentials.'),401

@app.get('/audit/view')
@audit_network_required()
def audit_view():
    if not session.get('audit_auth'): return redirect(url_for('audit_login'))
    return render_template('audit_view.html',events=AuditEvent.query.order_by(AuditEvent.id.desc()).limit(500).all())

def academic_percentage_map(student_ids):
    ids=[int(x) for x in student_ids if x is not None]
    if not ids: return {}
    rows=(db.session.query(Mark.student_id, func.coalesce(func.sum(Mark.marks),0), func.coalesce(func.sum(Mark.max_marks),0))
          .filter(Mark.student_id.in_(ids), Mark.marks.isnot(None))
          .group_by(Mark.student_id).all())
    return {sid: round((float(got)/float(mx))*100,2) if mx else 0 for sid,got,mx in rows}

@app.get('/ai')
@login_required()
def ai_page(): return render_template('ai.html',account=current_account())

@app.post('/api/ai')
@login_required()
def ai_api():
    question=(request.get_json(silent=True) or {}).get('message','').strip()
    if not question:
        return jsonify({'error':'Message required.'}),400

    acct=current_account(); start,end=school_year_bounds(); q=allowed_students_for_account(acct)
    # Keep database work bounded: one student query + one attendance aggregation query.
    students=q.order_by(Student.class_name,Student.section,Student.roll_number,Student.name).limit(1000).all()
    amap=attendance_percentage_map([st.id for st in students],start,min(school_date(),end))

    # Deterministic analytics answer for common attendance questions works even when no AI key is configured.
    m=re.search(r"(?:how many|number of)\s+(?:people|students)\s+(?:have|with)\s+(?:less than|below|under)\s+(\d+(?:\.\d+)?)\s*%?\s*(?:attendance)?", question.lower())
    if not m:
        m=re.search(r"(?:less than|below|under)\s+(\d+(?:\.\d+)?)\s*%?\s*attendance", question.lower())
    if m:
        threshold=float(m.group(1))
        matched=[st for st in students if amap.get(st.id,0) < threshold]
        return jsonify({'answer':f"{len(matched)} student{'s' if len(matched)!=1 else ''} in your allowed scope have attendance below {threshold:g}%.", 'source':'school-data'})

    # Deterministic academic analytics: no OpenAI key required for counts or lists.
    academic_pct=academic_percentage_map([st.id for st in students])
    academic_question = ('academic' in question.lower() or 'academics' in question.lower() or 'acedemic' in question.lower() or 'acedemics' in question.lower() or 'marks' in question.lower() or 'results' in question.lower())
    am=re.search(r"(?:less than|below|under)\s+(\d+(?:\.\d+)?)\s*%?", question.lower()) if academic_question else None
    if am:
        threshold=float(am.group(1)); matched=[st for st in students if academic_pct.get(st.id,0) < threshold]
        wants_list=bool(re.search(r"\b(which|who|list|students?)\b", question.lower()))
        names=', '.join(f"{st.name} ({st.class_name}{('-'+st.section) if st.section else ''}) — {academic_pct.get(st.id,0):g}%" for st in matched[:100])
        answer=f"{len(matched)} student{'s' if len(matched)!=1 else ''} in your allowed scope have academic performance below {threshold:g}%."
        if wants_list:
            answer += (f"\n{names}" if matched else "\nNo students match that academic filter.")
        return jsonify({'answer':answer,'source':'school-data'})

    # “Which students ... attendance?” is also deterministic and scope-safe.
    if re.search(r"\b(which|who|list)\b", question.lower()) and ('attendance' in question.lower() or 'absentee' in question.lower()):
        mm=re.search(r"(?:less than|below|under)\s+(\d+(?:\.\d+)?)", question.lower())
        if mm:
            threshold=float(mm.group(1)); matched=[st for st in students if amap.get(st.id,0) < threshold]
            names=', '.join(f"{st.name} ({st.class_name}{('-'+st.section) if st.section else ''}) — {amap.get(st.id,0):g}%" for st in matched[:50])
            return jsonify({'answer':(f"{len(matched)} student{'s' if len(matched)!=1 else ''} match your attendance filter.\n{names}" if matched else 'No students match that attendance filter.'),'source':'school-data'})

    if not os.getenv('OPENAI_API_KEY'):
        return jsonify({'answer':_local_ai_fallback(question,acct,students,amap,academic_pct),'source':'local-school-ai'})

    try:
        from openai import OpenAI
    except Exception:
        return jsonify({'error':'OpenAI package is not installed on this deployment.'}),503

    scope_ids=[st.id for st in students[:500]]
    invoice_rows=FeeInvoice.query.filter(FeeInvoice.student_id.in_(scope_ids)).order_by(FeeInvoice.created_at.desc()).all() if scope_ids else []
    inv_by_student={}
    for inv in invoice_rows:
        if len(inv_by_student.get(inv.student_id,[]))>=10: continue
        paid=sum(float(p.amount or 0) for p in inv.payments)
        inv_by_student.setdefault(inv.student_id,[]).append({'title':inv.title,'amount_due':float(inv.amount_due or 0),'status':inv.status,'due_date':inv.due_date.isoformat() if inv.due_date else None,'paid':paid})
    fee_rows=[{'student':st.name,'fees':inv_by_student.get(st.id,[])} for st in students[:500]]
    if acct.role=='parent':
        p=Parent.query.filter_by(account_id=acct.id).first(); parent_id=p.id if p else -1
        ann_rows=Announcement.query.filter(Announcement.published.is_(True), or_(Announcement.audience.in_(['public','all','parents']), Announcement.parent_id==parent_id)).order_by(Announcement.published_at.desc()).limit(30).all()
    else:
        ann_rows=Announcement.query.filter(Announcement.published.is_(True), Announcement.audience.in_(['public','all',{'admin':'admin','teacher':'teachers','student':'students','parent':'parents'}.get(acct.role)])).order_by(Announcement.published_at.desc()).limit(30).all()
    context={'role':acct.role,'school':'DAV PS KKP','students':[{
        'name':st.name,'class':st.class_name,'section':st.section,'admission':st.admission_number,
        'attendance':amap.get(st.id,0),'academic_percent':academic_pct.get(st.id,0)} for st in students[:500]],'fees':fee_rows,
        'announcements':[{'title':a.title,'message':a.message,'audience':a.audience,'published_at':a.published_at.isoformat() if a.published_at else None} for a in ann_rows]}
    if acct.role=='parent': context['instruction']='Only answer about children linked to this parent account.'
    elif acct.role=='student': context['instruction']='Only answer about the logged-in student and their own records.'
    elif acct.role=='teacher': context['instruction']='Only answer about students in the teacher assigned class/sections.'
    else: context['instruction']='Admin can answer school-wide questions from the supplied structured data.'

    instructions=("You are the DAV PS KKP School AI. Use only the structured school data provided in the user message. "
        "Never expose passwords, face encodings, private unrelated people, or data outside the logged-in user's scope. "
        "Do not invent marks, attendance, fees, or announcements. Be concise and practical.")
    try:
        timeout=float(os.getenv('OPENAI_TIMEOUT_SECONDS','15'))
        client=OpenAI(api_key=os.getenv('OPENAI_API_KEY'),timeout=timeout)
        response=client.responses.create(model=os.getenv('OPENAI_MODEL','gpt-5.6'),reasoning={'effort':os.getenv('OPENAI_REASONING_EFFORT','low')},text={'verbosity':'low'},instructions=instructions,
            input=f"STRUCTURED DATA:\n{json.dumps(context,ensure_ascii=False)}\n\nUSER QUESTION:\n{question}")
        return jsonify({'answer':response.output_text,'source':'openai'})
    except Exception as exc:
        app.logger.exception('OpenAI request failed')
        # Keep the school-data assistant useful even if OpenAI is rate-limited/unavailable.
        if 'academic' in question.lower() or 'academics' in question.lower() or 'marks' in question.lower() or 'results' in question.lower():
            am=re.search(r"(?:less than|below|under)\s+(\d+(?:\.\d+)?)\s*%?", question.lower())
            if am:
                threshold=float(am.group(1)); matched=[st for st in students if academic_pct.get(st.id,0) < threshold]
                names=', '.join(f"{st.name} ({st.class_name}{('-'+st.section) if st.section else ''}) — {academic_pct.get(st.id,0):g}%" for st in matched[:100])
                answer=f"{len(matched)} student{'s' if len(matched)!=1 else ''} in your allowed scope have academic performance below {threshold:g}%."
                if re.search(r"\b(which|who|list|students?)\b", question.lower()):
                    answer += (f"\n{names}" if matched else "\nNo students match that academic filter.")
                return jsonify({'answer':answer,'source':'school-data-fallback'})
        return jsonify({'answer':_local_ai_fallback(question,acct,students,amap,academic_pct),'source':'local-school-ai-fallback'})

@app.get('/announcements')
def announcements_view():
    acct=current_account() if session.get('account_id') else None
    if not acct:
        announcements=Announcement.query.filter_by(published=True,audience='public').order_by(Announcement.published_at.desc()).limit(30).all()
        return render_template('announcements_public.html',announcements=announcements,logged_in=False,role='public')
    audience={'admin':'admin','teacher':'teachers','student':'students','parent':'parents'}.get(acct.role)
    if acct.role=='parent':
        p=Parent.query.filter_by(account_id=acct.id).first()
        announcements=Announcement.query.filter(
            Announcement.published.is_(True),
            or_(Announcement.audience.in_(['public','all','parents']), Announcement.parent_id== (p.id if p else -1))
        ).order_by(Announcement.published_at.desc()).limit(50).all()
    else:
        announcements=Announcement.query.filter(
            Announcement.published.is_(True),
            Announcement.audience.in_(['public','all',audience])
        ).order_by(Announcement.published_at.desc()).limit(50).all()
    return render_template('announcements_public.html',announcements=announcements,logged_in=True,role=acct.role)


@app.post('/api/auth/login')
def api_auth_login():
    data=request.get_json(silent=True) or {}; username=str(data.get('username','')).strip(); password=str(data.get('password',''))
    acct=Account.query.filter(func.lower(Account.username)==username.lower()).first()
    if not acct or not acct.active or not check_password_hash(acct.password_hash,password): return jsonify({'error':'Invalid username or password.'}),401
    try: token=_issue_jwt(acct)
    except Exception as exc: return jsonify({'error':f'JWT support unavailable: {exc}'}),503
    return jsonify({'access_token':token,'token_type':'Bearer','expires_in':app.config['JWT_ACCESS_TTL_MINUTES']*60,'account':{'id':acct.id,'username':acct.username,'role':acct.role,'display_name':acct.display_name}})

@app.get('/api/auth/me')
@login_required()
def api_auth_me():
    acct=current_account(); return jsonify({'id':acct.id,'username':acct.username,'role':acct.role,'display_name':acct.display_name})

@app.route('/register', methods=['GET','POST'])
def student_self_register():
    if request.method=='POST':
        d=request.form; name=normalize_school_name(d.get('name')); adm=str(d.get('admission_number','')).strip().upper(); roll=str(d.get('roll_number','')).strip().upper(); cls=normalize_class(d.get('class_name')); sec=normalize_section(d.get('section')); username=str(d.get('username','')).strip(); password=d.get('password',''); second=(d.get('second_language') or '').strip().lower() or None; third=(d.get('third_language') or '').strip().lower() or None
        if class_number(cls)>=9: third=None
        section_required=class_number(cls) not in {11,12}
        if not all([name,adm,roll,cls,username]) or (section_required and not sec) or len(password)<8: return render_template('register_self_student.html',class_options=class_list(),form=d,error='Fill all required fields. Section is not required for XI/XII.'),400
        if Student.query.filter_by(admission_number=adm).first(): return render_template('register_self_student.html',class_options=class_list(),form=d,error='Admission number already exists.'),409
        if Account.query.filter(func.lower(Account.username)==username.lower()).first(): return render_template('register_self_student.html',class_options=class_list(),form=d,error='Username already exists.'),409
        if class_number(cls) in range(5,9) and (not second or not third): return render_template('register_self_student.html',class_options=class_list(),form=d,error='2nd and 3rd languages are required for V–VIII.'),400
        if class_number(cls) in range(9,11) and not second: return render_template('register_self_student.html',class_options=class_list(),form=d,error='2nd language is required for IX–X.'),400
        acct=Account(username=username,password_hash=generate_password_hash(password),role='student',display_name=name,must_change_password=False)
        db.session.add(acct); db.session.flush(); st=Student(name=name,admission_number=adm,roll_number=roll,class_name=cls,section=sec,second_language=second,third_language=third,phone=d.get('phone','').strip(),email=d.get('email','').strip(),account_id=acct.id)
        db.session.add(st); log_audit('student_self_registered','Student',st.id,{'class':cls,'section':sec,'username':username}); db.session.commit()
        return redirect(url_for('login',registered='1'))
    return render_template('register_self_student.html',class_options=class_list(),form={})


def _admin_user_or_404(model, item_id):
    item=db.session.get(model,item_id)
    if not item: abort(404)
    return item

@app.route('/admin/teachers/<int:tid>/edit', methods=['GET','POST'])
@admin_required
def edit_teacher(tid):
    t=_admin_user_or_404(Teacher,tid); acct=t.account
    if request.method=='POST':
        name=normalize_school_name(request.form.get('name')); email=request.form.get('email','').strip(); phone=request.form.get('phone','').strip(); username=request.form.get('username','').strip(); active=request.form.get('active')=='1'
        if not name or not username: flash('Name and username are required.','error')
        elif acct is None: flash('Teacher account is missing. Please recreate or migrate this teacher account first.','error')
        elif username.lower()!=acct.username.lower() and Account.query.filter(func.lower(Account.username)==username.lower()).first(): flash('Username already exists.','error')
        else:
            t.name=name; t.email=email; t.phone=phone; t.active=active; acct.display_name=name; acct.username=username; acct.active=active
            # Legacy subject assignments are retained in the database for backwards
            # compatibility, but they are no longer used or required.
            log_audit('teacher_updated','Teacher',tid,{'mode':'class_teacher'}); db.session.commit(); flash('TEACHER UPDATED.','success'); return redirect(url_for('admin_accounts',role='teacher'))
    return render_template('person_edit.html',kind='Teacher',person=t,account=acct)


@app.post('/admin/teachers/<int:tid>/delete')
@admin_required
def delete_teacher(tid):
    t=_admin_user_or_404(Teacher,tid); acct=t.account
    if acct and acct.id==current_account().id: abort(400)
    TeacherAssignment.query.filter_by(teacher_id=t.id).delete(synchronize_session=False)
    if acct: db.session.delete(acct)
    db.session.delete(t); log_audit('teacher_deleted','Teacher',tid); db.session.commit(); flash('Teacher deleted.','success'); return redirect(url_for('admin_accounts',role='teacher'))

@app.route('/admin/students/<int:sid>/edit', methods=['GET','POST'])
@admin_required
def edit_student(sid):
    st=_admin_user_or_404(Student,sid)
    acct=st.account
    parents=Parent.query.order_by(Parent.name).all()
    linked={x.parent_id for x in ParentStudent.query.filter_by(student_id=sid).all()}
    if request.method=='POST':
        name=normalize_school_name(request.form.get('name')); cls=normalize_class(request.form.get('class_name')); sec=normalize_section(request.form.get('section')); roll=request.form.get('roll_number','').strip().upper(); username=request.form.get('username','').strip(); active=request.form.get('active')=='1'; temp_password=request.form.get('temp_password','')
        if class_number(cls) in {11,12}: sec=''
        if not name or not cls or (class_number(cls) not in {11,12} and not sec) or not username:
            flash('Name, class, section (except XI/XII) and username are required.','error')
        elif (acct is None and len(temp_password)<8):
            flash('This legacy student is missing an account. Enter an 8+ character temporary password to create it.','error')
        elif acct is not None and username.lower()!=acct.username.lower() and Account.query.filter(func.lower(Account.username)==username.lower()).first():
            flash('Username already exists.','error')
        elif acct is None and Account.query.filter(func.lower(Account.username)==username.lower()).first():
            flash('Username already exists.','error')
        else:
            st.name=name; st.class_name=cls; st.section=sec; st.roll_number=roll; st.active=active
            if acct is None:
                acct=Account(username=username,password_hash=generate_password_hash(temp_password),role='student',display_name=name,must_change_password=True,active=active)
                db.session.add(acct); db.session.flush(); st.account_id=acct.id
            else:
                acct.display_name=name; acct.username=username; acct.active=active
                if temp_password:
                    if len(temp_password)<8: flash('Temporary password must be at least 8 characters.','error'); return render_template('student_edit.html',student=st,account=acct,parents=parents,linked_parents=linked,class_options=class_list())
                    acct.password_hash=generate_password_hash(temp_password); acct.must_change_password=True
            ParentStudent.query.filter_by(student_id=sid).delete(synchronize_session=False)
            for pid in [int(x) for x in request.form.getlist('parent_ids') if str(x).isdigit()]:
                if db.session.get(Parent,pid): db.session.add(ParentStudent(parent_id=pid,student_id=sid))
            log_audit('student_updated','Student',sid); db.session.commit(); flash('Student updated.','success'); return redirect(url_for('admin_students',class_name=cls,section=sec))
    return render_template('student_edit.html',student=st,account=acct,parents=parents,linked_parents=linked,class_options=class_list())

@app.post('/admin/students/<int:sid>/delete')
@admin_required
def delete_student(sid):
    st=_admin_user_or_404(Student,sid); acct=st.account
    ParentStudent.query.filter_by(student_id=sid).delete(synchronize_session=False)
    for model,col in [(Attendance,Attendance.student_id),(Mark,Mark.student_id),(AssessmentComponent,AssessmentComponent.student_id),(ReportCardConfig,ReportCardConfig.student_id),(PublishedReport,PublishedReport.student_id)]:
        model.query.filter(col==sid).delete(synchronize_session=False)
    invoices=FeeInvoice.query.filter_by(student_id=sid).all(); [FeePayment.query.filter_by(invoice_id=i.id).delete(synchronize_session=False) for i in invoices]; FeeInvoice.query.filter_by(student_id=sid).delete(synchronize_session=False)
    if acct: db.session.delete(acct)
    db.session.delete(st); log_audit('student_deleted','Student',sid); db.session.commit(); flash('Student deleted.','success'); return redirect(url_for('admin_students'))

@app.route('/admin/parents/<int:pid>/edit', methods=['GET','POST'])
@admin_required
def edit_parent(pid):
    parent=_admin_user_or_404(Parent,pid); acct=parent.account; students=Student.query.filter_by(active=True).order_by(Student.class_name,Student.section,Student.name).all(); linked={x.student_id for x in ParentStudent.query.filter_by(parent_id=pid).all()}
    if request.method=='POST':
        name=normalize_school_name(request.form.get('name')); username=request.form.get('username','').strip(); admissions=[x.strip().upper() for x in request.form.get('child_admissions','').replace('\n',',').split(',') if x.strip()]; ids=[s.id for s in Student.query.filter(Student.admission_number.in_(admissions)).all()]
        if not name or not username or not ids: flash('Parent name, username and at least one linked student are required.','error')
        elif username.lower()!=acct.username.lower() and Account.query.filter(func.lower(Account.username)==username.lower()).first(): flash('Username already exists.','error')
        else:
            parent.name=name; parent.phone=request.form.get('phone','').strip(); parent.email=request.form.get('email','').strip(); acct.display_name=name; acct.username=username; acct.active=request.form.get('active')=='1'
            ParentStudent.query.filter_by(parent_id=pid).delete(synchronize_session=False)
            for sid in ids:
                if db.session.get(Student,sid): db.session.add(ParentStudent(parent_id=pid,student_id=sid))
            log_audit('parent_updated','Parent',pid,{'children':ids}); db.session.commit(); flash('Parent updated.','success'); return redirect(url_for('admin_accounts',role='parent'))
    return render_template('parent_edit.html',parent=parent,account=acct,students=students,linked_students=linked,linked_admissions=[s.admission_number for s in students if s.id in linked])

@app.post('/admin/parents/<int:pid>/delete')
@admin_required
def delete_parent(pid):
    parent=_admin_user_or_404(Parent,pid); acct=parent.account
    ParentStudent.query.filter_by(parent_id=pid).delete(synchronize_session=False)
    if acct: db.session.delete(acct)
    db.session.delete(parent); log_audit('parent_deleted','Parent',pid); db.session.commit(); flash('Parent deleted.','success'); return redirect(url_for('admin_accounts',role='parent'))

@app.post('/admin/assignments/<int:aid>/delete')
@admin_required
def delete_assignment(aid):
    a=_admin_user_or_404(TeacherAssignment,aid); db.session.delete(a); log_audit('teacher_assignment_deleted','TeacherAssignment',aid); db.session.commit(); flash('Class teacher assignment removed.','success'); return redirect(url_for('assignments'))

@app.route('/admin/assignments/<int:aid>/edit', methods=['GET','POST'])
@admin_required
def edit_assignment(aid):
    a=_admin_user_or_404(TeacherAssignment,aid); teachers=Teacher.query.filter_by(active=True).order_by(Teacher.name).all()
    if request.method=='POST':
        tid=int(request.form.get('teacher_id')); cls=normalize_class(request.form.get('class_name')); sec=normalize_section(request.form.get('section'))
        conflict=TeacherAssignment.query.filter(TeacherAssignment.id!=aid,TeacherAssignment.class_name==cls,TeacherAssignment.section==sec).first()
        if conflict: flash('That class/section already has a class teacher.','error')
        else: a.teacher_id=tid; a.class_name=cls; a.section='' if class_number(cls) in {11,12} else sec; log_audit('teacher_assignment_updated','TeacherAssignment',aid,{'teacher_id':tid}); db.session.commit(); flash('Assignment updated.','success'); return redirect(url_for('assignments'))
    return render_template('assignment_edit.html',assignment=a,teachers=teachers,class_options=class_list())

@app.route('/admin/test-clock', methods=['GET','POST'])
@admin_required
def test_clock():
    row=db.session.get(SchoolClock,1)
    if not row:
        row=SchoolClock(id=1,override_time=None)
        db.session.add(row)
    if request.method=='POST':
        raw=str(request.form.get('override_time','')).strip()
        action=str(request.form.get('action','save')).strip().lower()
        if action=='clear' or raw.lower() in {'','real','system'}:
            row.override_time=None
            log_audit('school_clock_cleared','SchoolClock',1)
            db.session.commit()
            flash('TEST CLOCK CLEARED. REAL SCHOOL TIME IS ACTIVE.','success')
        else:
            try:
                value=datetime.strptime(raw,'%H:%M').strftime('%H:%M')
            except ValueError:
                flash('ENTER TIME IN 24-HOUR FORMAT, FOR EXAMPLE 08:45.','error')
                return render_template('test_clock.html',clock=row,school_time=school_time(),status=attendance_status_for_time())
            row.override_time=value
            row.updated_at=datetime.utcnow()
            log_audit('school_clock_updated','SchoolClock',1,{'override_time':value})
            db.session.commit()
            flash(f'TEST CLOCK SET TO {value}. ATTENDANCE USES THIS TIME.','success')
    return render_template('test_clock.html',clock=row,school_time=school_time(),status=attendance_status_for_time())

@app.errorhandler(500)
def internal_server_error(e):
    try:
        db.session.rollback()
    except Exception:
        pass
    app.logger.exception('Unhandled HTTP 500')
    return render_template('error.html',code=500,message='Something went wrong on the server. Your data was rolled back. Please retry once; if it repeats, check Render logs.'),500

@app.get('/settings')
@login_required()
def settings(): return redirect(url_for('change_credentials'))

@app.errorhandler(404)
def not_found(e): return render_template('error.html',code=404,message='Page not found.'),404
@app.errorhandler(403)
def forbidden(e): return render_template('error.html',code=403,message='You do not have access to this area.'),403

if __name__=='__main__':
    app.run(host='0.0.0.0',port=int(os.getenv('PORT','5000')),debug=True)
